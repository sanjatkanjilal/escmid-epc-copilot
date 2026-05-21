#!/usr/bin/env python3
"""
ESCMID Proposal Reviewer
========================
Reads ESCMID proposal HTML files, applies the 94-tag clinical taxonomy,
scores each proposal against the historical programme (from sessions_raw.json),
and generates an interactive HTML reviewer dashboard.

Quick start
-----------
1.  Run the programme analyser first (or use existing sessions_raw.json):
        python escmid_analyser.py --rebuild

2.  Place proposal HTML files in  data/proposals/
        (save each proposal page in Chrome as "View Source", then save as .html)

3.  Run:
        python proposal_reviewer.py               # tag + score + dashboard
        python proposal_reviewer.py --skip-tagging  # use cache / keyword fallback

Outputs
-------
    data/output/proposals_tagged.json       intermediate (safe to delete)
    data/output/Proposal_Review.html        interactive dashboard
    data/output/Proposal_Review.xlsx        Excel with tags + scores for annotation

"""

import argparse, json, os, re, sys, time
from collections import defaultdict, Counter
from pathlib import Path

# ── Optional deps ─────────────────────────────────────────────────────────────
try:
    from dotenv import load_dotenv; load_dotenv()
except ImportError:
    pass

try:
    import anthropic as _anthropic
    HAS_ANTHROPIC = True
except ImportError:
    HAS_ANTHROPIC = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from tqdm import tqdm as _tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False
    _tqdm = lambda x, **_: x

# ── Import proposal parser ────────────────────────────────────────────────────
_SCRIPT_DIR = Path(__file__).parent
sys.path.insert(0, str(_SCRIPT_DIR))
try:
    import parse_proposals_vscode as _parser
    HAS_PARSER = True
except ImportError:
    HAS_PARSER = False
    print("Warning: parse_proposals_vscode.py not found — cannot parse HTML files")

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR       = _SCRIPT_DIR
DATA_DIR       = BASE_DIR / "data"
PROPOSALS_DIR  = DATA_DIR / "proposals"
OUTPUT_DIR     = DATA_DIR / "output"
SESSIONS_JSON  = OUTPUT_DIR / "sessions_raw.json"
CACHE_FILE     = DATA_DIR / "tagging_cache.json"

ANTHROPIC_MODEL = "claude-sonnet-4-5"
API_DELAY       = 0.4
MAX_RETRIES     = 3
SAVE_EVERY      = 25

# ── Load 94-tag taxonomy from escmid_analyser if available ────────────────────
try:
    import escmid_analyser as _ea
    ALL_TAGS    = _ea.ALL_TAGS
    TAG_PALETTE = _ea.TAG_PALETTE
    TAG_SET     = _ea.TAG_SET
    TAG_TAXONOMY= _ea.TAG_TAXONOMY
except Exception:
    # Fallback: define inline (keep in sync with escmid_analyser.py)
    TAG_TAXONOMY = {
        "Methods":    ["methods_Prediction modeling","methods_Large language models",
                       "methods_Image recognition","methods_Causal inference",
                       "methods_Genomics and sequencing",
                       "methods_Systematic review and meta-analysis",
                       "methods_Pharmacokinetics and pharmacodynamics",
                       "methods_Epidemiological modelling"],
        "Study Design":["study_design_RCT","study_design_Cohort","study_design_Case series",
                        "study_design_Modelling","study_design_Basic science","study_design_Review"],
        "ClinMicro":  ["clinmicro_Bacteriology","clinmicro_Mycobacteriology","clinmicro_Mycology",
                       "clinmicro_Virology","clinmicro_Parasitology","clinmicro_Prions",
                       "clinmicro_Diagnostics","clinmicro_Antimicrobial resistance",
                       "clinmicro_Diagnostic stewardship","clinmicro_Basic science"],
        "Infectious Diseases":["infectiousdz_Antimicrobial stewardship",
                       "infectiousdz_Infection prevention and control",
                       "infectiousdz_Treatment","infectiousdz_Comparative trials",
                       "infectiousdz_Emerging infections","infectiousdz_Epidemiology",
                       "infectiousdz_Clinical manifestations","infectiousdz_Case series"],
        "Treatments": ["treatments_Antibiotics","treatments_Antivirals","treatments_Antifungals",
                       "treatments_Antiparasitics","treatments_Vaccines","treatments_Phage therapy",
                       "treatments_Immunotherapy","treatments_Experimental therapy"],
        "Syndromes":  ["syndrome_Fever of unknown origin","syndrome_Febrile neutropaenia",
                       "syndrome_Upper respiratory tract infection",
                       "syndrome_Lower respiratory tract infection",
                       "syndrome_Urinary tract infection",
                       "syndrome_Sepsis and bloodstream infection",
                       "syndrome_Endocarditis and cardiovascular infection",
                       "syndrome_Intraabdominal infection","syndrome_Hepatitis and liver infection",
                       "syndrome_Gastrointestinal infection","syndrome_Meningitis and encephalitis",
                       "syndrome_Brain abscess","syndrome_Skin and soft tissue infection",
                       "syndrome_Bone and joint infection","syndrome_Ocular infection",
                       "syndrome_Sexually transmitted infection",
                       "syndrome_HIV and AIDS","syndrome_Surgical site infection"],
        "Special Hosts":["hosts_Immunocompromised","hosts_Transplant",
                       "hosts_Oncology and haematology","hosts_Paediatric","hosts_Neonatal",
                       "hosts_Elderly","hosts_Pregnancy and maternal","hosts_ICU and critically ill"],
        "AMR Pathogens":["amr_pathogen_ESKAPE","amr_pathogen_MRSA","amr_pathogen_C. difficile",
                       "amr_pathogen_Mycobacteria","amr_pathogen_Gonorrhoea","amr_pathogen_Candida"],
        "Microbiome":  ["microbiome_Gut","microbiome_Respiratory","microbiome_Skin",
                        "microbiome_Intervention"],
        "Public Health":["public_health_Surveillance","public_health_Pandemic preparedness",
                        "public_health_Policy","public_health_One Health and zoonoses",
                        "public_health_Travel medicine","public_health_Outbreak response",
                        "public_health_Bioterrorism and biosecurity"],
        "Professional":["professional_Regulatory","professional_Ethics",
                        "professional_Bias and equity","professional_Education and training",
                        "professional_Guidelines","professional_Career development"],
        "Region":      ["region_LMICs","region_Europe","region_Africa",
                        "region_Asia-Pacific","region_Americas"],
    }
    ALL_TAGS    = [t for g in TAG_TAXONOMY.values() for t in g]
    TAG_SET     = set(ALL_TAGS)
    TAG_PALETTE = {
        "methods_":"#22d3ee","study_design_":"#818cf8","clinmicro_":"#fbbf24",
        "infectiousdz_":"#34d399","treatments_":"#fb923c","syndrome_":"#f472b6",
        "hosts_":"#a78bfa","amr_pathogen_":"#f87171","microbiome_":"#6ee7b7",
        "public_health_":"#86efac","professional_":"#94a3b8","region_":"#4ade80",
    }

def tag_color(tag):
    for prefix, color in TAG_PALETTE.items():
        if tag.startswith(prefix): return color
    return "#617d9b"

ALL_TAGS_STR = "\n".join(f"  {t}" for t in ALL_TAGS)

# ══════════════════════════════════════════════════════════════════════════════
# PROPOSAL TAGGING
# ══════════════════════════════════════════════════════════════════════════════

TAG_SYSTEM = (
    "You are an expert in infectious diseases and clinical microbiology. "
    "Tag ESCMID conference session proposals using a fixed clinical taxonomy. "
    "Respond ONLY with valid JSON, no preamble or markdown."
)

def build_proposal_tag_prompt(proposal: dict) -> str:
    title   = proposal.get("session_title", "")
    topics  = proposal.get("topic_titles", "")[:400]
    motiv   = proposal.get("motivation", "")[:300]
    stype   = proposal.get("session_type", "")
    cat     = proposal.get("category", "")

    return f"""Assign 1-4 tags to this ESCMID conference session proposal.

Session Title: {title}
Session Type:  {stype}
Category:      {cat}
Topics:        {topics}
Motivation:    {motiv}

Taxonomy:
{ALL_TAGS_STR}

Priority: syndrome → hosts → amr_pathogen → clinmicro → treatments → infectiousdz → methods → public_health

Return JSON only: {{"tags": ["tag1", "tag2"]}}
Rules: max 4, min 1, exact tag names from list above."""


def load_cache() -> dict:
    return json.loads(CACHE_FILE.read_text()) if CACHE_FILE.exists() else {}

def save_cache(cache: dict):
    CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
    CACHE_FILE.write_text(json.dumps(cache, indent=2))


def tag_proposal_api(proposal: dict, client, cache: dict) -> list:
    pid = proposal.get("proposal_id", "")
    key = f"proposal_tag|{pid}|{proposal.get('session_title','')[:50]}"
    if key in cache:
        return cache[key]

    for attempt in range(MAX_RETRIES):
        try:
            resp = client.messages.create(
                model=ANTHROPIC_MODEL, max_tokens=150,
                system=TAG_SYSTEM,
                messages=[{"role":"user","content":build_proposal_tag_prompt(proposal)}]
            )
            if not resp.content: raise ValueError("empty")
            raw = re.sub(r'^```json?\s*|\s*```$', '', resp.content[0].text.strip())
            tags = [t for t in json.loads(raw).get("tags",[]) if t in TAG_SET][:4]
            cache[key] = tags
            return tags
        except Exception:
            if attempt == MAX_RETRIES - 1: break
            time.sleep(1.0)
    cache[key] = []
    return []


def keyword_tag_proposal(proposal: dict) -> list:
    """Simple keyword fallback — mirrors escmid_analyser keyword rules."""
    title = (proposal.get("session_title","") + " " +
             proposal.get("topic_titles","") + " " +
             proposal.get("motivation","")).lower()
    rules = [
        ("syndrome_Sepsis and bloodstream infection", ["sepsis","bacteraemia","bacteremia"]),
        ("syndrome_Lower respiratory tract infection", ["pneumonia","respiratory infection"]),
        ("syndrome_Urinary tract infection",           ["urinary tract","uti "]),
        ("clinmicro_Antimicrobial resistance",         ["resistance","amr","mdr","carbapenem"]),
        ("infectiousdz_Antimicrobial stewardship",     ["stewardship","prescribing"]),
        ("methods_Large language models",              ["artificial intelligence","machine learning","llm","ai "]),
        ("clinmicro_Mycology",                         ["fungal","candida","aspergillus","antifungal"]),
        ("clinmicro_Virology",                         ["viral","virus","hiv","covid","hepatitis"]),
        ("public_health_Surveillance",                 ["surveillance","epidemiology"]),
        ("treatments_Vaccines",                        ["vaccine","vaccination"]),
        ("hosts_Immunocompromised",                    ["immunocompromised","transplant","haematology"]),
        ("clinmicro_Diagnostics",                      ["diagnostic","diagnosis","test"]),
    ]
    tags = []
    for tag, keywords in rules:
        if any(k in title for k in keywords):
            tags.append(tag)
            if len(tags) >= 4: break
    return tags or ["clinmicro_Diagnostics"]


def tag_all_proposals(proposals: list, use_api: bool) -> list:
    cache  = load_cache()
    client = None
    if use_api:
        if not HAS_ANTHROPIC:
            print("  anthropic not installed — using keyword fallback")
            use_api = False
        else:
            api_key = os.environ.get("ANTHROPIC_API_KEY","")
            if not api_key:
                print("  ANTHROPIC_API_KEY not set — using keyword fallback")
                use_api = False
            else:
                client = _anthropic.Anthropic(api_key=api_key)

    it = _tqdm(proposals, desc="  Tagging proposals") if HAS_TQDM else proposals
    for i, p in enumerate(it):
        if not p.get("tags"):
            p["tags"] = tag_proposal_api(p, client, cache) if use_api \
                        else keyword_tag_proposal(p)
            if use_api:
                if i % SAVE_EVERY == 0: save_cache(cache)
                time.sleep(API_DELAY)
    save_cache(cache)
    return proposals


# ══════════════════════════════════════════════════════════════════════════════
# SCORING
# ══════════════════════════════════════════════════════════════════════════════

def jaccard(a: set, b: set) -> float:
    if not a or not b: return 0.0
    return len(a & b) / len(a | b)


def compute_ai_rating(proposal: dict) -> dict:
    """
    Heuristic AI ratings from available data (no API call).
    C1-C8 scored 1-5; C9/C10/C11 = None (insufficient data).
    """
    tags    = set(proposal.get("tags") or [])
    novelty = float(proposal.get("score_novelty") or 0.5)
    trend   = float(proposal.get("score_trend")   or 0)
    motiv   = proposal.get("motivation") or ""
    ntopics = int(proposal.get("num_topics")  or 0)
    stype   = (proposal.get("session_type")   or "").lower()
    prop    = proposal.get("proposing_entities") or ""

    c1 = max(1, min(5, round(3 + trend * 1.5 + (novelty - 0.5))))
    c2 = max(1, min(5, round(1 + novelty * 4)))
    groups = set(t.split("_")[0] + "_" for t in tags)
    c3 = max(1, min(5, len(groups) + 1))
    has_basic = any(t in ("study_design_Basic science", "clinmicro_Basic science") for t in tags)
    has_clin  = any(t.startswith(p) for t in tags
                    for p in ("syndrome_", "infectiousdz_Treatment", "hosts_", "treatments_"))
    c4 = 5 if has_basic and has_clin else 4 if has_clin else 3 if has_basic else 2
    c5 = 3
    c6 = 4 if len(prop) > 15 else 3
    expected = {"keynote":1,"interview":1,"symposium":3,"educational":4,
                "workshop":4,"meet":1,"oral":6,"pipeline":3,"late":3}
    exp = next((v for k, v in expected.items() if k in stype), 3)
    c7 = 4 if ntopics > 0 and abs(ntopics - exp) <= 1 else 3 if ntopics > 0 else 2
    mlen = len(motiv)
    c8 = 5 if mlen > 500 else 4 if mlen > 250 else 3 if mlen > 100 else 2 if mlen > 20 else 1
    overall = round((c1 + c2 + c3 + c4 + c5 + c6 + c7 + c8) / 8)
    rationale = (
        f"Novelty {novelty:.2f} ({'new topic' if novelty > 0.6 else 'previously covered'}), "
        f"trend {'↑ growing' if trend > 0.1 else '↓ declining' if trend < -0.1 else '→ stable'}, "
        f"{len(groups)} tag domains, "
        f"description {'detailed' if mlen > 300 else 'brief' if mlen > 50 else 'minimal'}."
    )
    return {"C1":c1,"C2":c2,"C3":c3,"C4":c4,"C5":c5,"C6":c6,"C7":c7,"C8":c8,
            "C9":None,"C10":None,"C11":None,"overall":overall,"rationale":rationale}


def api_rate_proposal(proposal: dict, client, cache: dict) -> dict:
    """Full Claude evaluation on all 11 EPC criteria. Cached by proposal_id."""
    pid = proposal.get("proposal_id", "")
    key = f"ai_review|{pid}"
    if key in cache:
        return cache[key]

    prompt = f"""Rate this ESCMID 2027 session proposal on 11 EPC criteria. Be concise and evidence-based.

Title:       {proposal.get("session_title","")}
Category:    {proposal.get("category","")}
Type:        {proposal.get("session_type","")}
Chairs:      {(proposal.get("chairs") or "")[:150]}
Topics:      {(proposal.get("topic_titles") or "")[:300]}
Description: {(proposal.get("motivation") or "")[:500]}
Tags:        {", ".join(proposal.get("tags") or [])}
Novelty:     {proposal.get("score_novelty","?")}  (0=seen before, 1=new topic)
Trend:       {proposal.get("score_trend","?")}    (positive=growing in programme)

Score each criterion 1(poor)–5(excellent) or "N/A" if not assessable from the info above.
C1: Hot/timely/controversial
C2: Not duplicated from recent meetings (use novelty score)
C3: Cross-disciplinary/wide appeal
C4: Basic+translational+clinical integration
C5: Appropriate session format
C6: Relevant collaborators involved
C7: Adheres to session format rules
C8: Proper description provided
C9: Gender/geographic balance of speakers
C10: Best speakers, not self-serving
C11: Engages young investigators

Return JSON only — no preamble:
{{"C1":3,"C2":4,"C3":3,"C4":2,"C5":3,"C6":"N/A","C7":3,"C8":4,"C9":"N/A","C10":"N/A","C11":"N/A","overall":3,"rationale":"One concise sentence."}}"""

    for attempt in range(MAX_RETRIES):
        try:
            resp = client.messages.create(
                model=ANTHROPIC_MODEL, max_tokens=250,
                messages=[{"role": "user", "content": prompt}]
            )
            raw = re.sub(r"^```json?\s*|\s*```$", "", resp.content[0].text.strip())
            result = json.loads(raw)
            cache[key] = result
            return result
        except Exception:
            if attempt == MAX_RETRIES - 1:
                break
            time.sleep(1.0)

    fallback = {c: "N/A" for c in ["C1","C2","C3","C4","C5","C6","C7","C8","C9","C10","C11"]}
    fallback.update({"overall": "N/A", "rationale": "Rating unavailable."})
    cache[key] = fallback
    return fallback


def rate_all_proposals(proposals: list, use_api: bool) -> list:
    """Add ai_rating to every proposal."""
    cache  = load_cache()
    client = None

    if use_api:
        if not HAS_ANTHROPIC:
            print("  anthropic not installed — using heuristic ratings")
            use_api = False
        else:
            api_key = os.environ.get("ANTHROPIC_API_KEY", "")
            if not api_key:
                print("  ANTHROPIC_API_KEY not set — using heuristic ratings")
                use_api = False
            else:
                client = _anthropic.Anthropic(api_key=api_key)
                print(f"  Mode: full Claude evaluation ({len(proposals)} proposals, ~$1–2)")

    it = _tqdm(proposals, desc="  Rating proposals") if HAS_TQDM else proposals
    for i, p in enumerate(it):
        if not p.get("ai_rating"):
            if use_api and client:
                p["ai_rating"] = api_rate_proposal(p, client, cache)
                if i % SAVE_EVERY == 0:
                    save_cache(cache)
                time.sleep(API_DELAY)
            else:
                p["ai_rating"] = compute_ai_rating(p)

    if use_api:
        save_cache(cache)
    return proposals


def score_proposals(proposals: list, hist_sessions: list) -> list:
    """
    Compute three scores for each proposal:

    novelty    (0-1)  1 = completely new topic, 0 = duplicate of past session
    trend      (-1,1) positive = topic growing 2022→2026, negative = shrinking
    gap_fill   (0-1)  1 = addresses subcategory with 0 historical sessions
    """
    YEARS = sorted(set(s.get("year","") for s in hist_sessions))
    early = set(y for y in YEARS if y <= "2023")
    late  = set(y for y in YEARS if y >= "2024")

    # Pre-compute tag sets per session
    hist = [(s, set(s.get("tags",[]))) for s in hist_sessions]

    # Subcategory counts for gap analysis
    subcat_counts = Counter(s.get("escmid_subcat","") for s in hist_sessions
                            if s.get("escmid_subcat"))

    # Similar sessions cache: proposal_id -> [(score, session)]
    similar_cache: dict = {}

    for p in proposals:
        ptags = set(p.get("tags",[]))

        # ── Novelty: max similarity to any historical session
        sims = [(jaccard(ptags, htags), s) for s, htags in hist if htags]
        sims.sort(key=lambda x: -x[0])
        max_sim = sims[0][0] if sims else 0.0
        p["score_novelty"] = round(1 - max_sim, 3)
        p["similar_sessions"] = [
            {"code": s["code"], "year": s["year"], "title": s["title"][:80],
             "similarity": round(sc, 3), "tags": s.get("tags",[])}
            for sc, s in sims[:5] if sc > 0
        ]

        # ── Trend: compare tag frequency in early vs late years
        early_count = sum(
            1 for s, htags in hist
            if s.get("year","") in early and ptags & htags
        )
        late_count  = sum(
            1 for s, htags in hist
            if s.get("year","") in late and ptags & htags
        )
        n_early = max(len(early), 1)
        n_late  = max(len(late), 1)
        trend = (late_count / n_late) - (early_count / n_early)
        # Normalise to [-1, 1] range roughly
        p["score_trend"]   = round(max(-1.0, min(1.0, trend * 2)), 3)
        p["trend_detail"]  = {"early": early_count, "late": late_count}

        # ── Gap fill: does it address a rare subcategory?
        pcat = p.get("category","")
        # Try to match proposal category to ESCMID subcategory counts
        # Use the category string from the proposal form as a proxy
        p["score_gap"] = 0.0  # enriched below if subcategory info available

    return proposals


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL OUTPUT
# ══════════════════════════════════════════════════════════════════════════════

def build_excel(proposals: list, path: Path):
    if not HAS_OPENPYXL:
        print("  openpyxl not installed — skipping Excel")
        return
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    CRITERIA = [('C1', 'Hot / Timely / Controversial', 'Is the topic hot, timely, or controversial? Does it reflect current scientific priorities and recent developments?'), ('C2', 'Not Duplicated from Recent Meetings', 'Not duplicated from ESCMID Global 2024/2025/2026, unless new developments or continued education justify resubmission.'), ('C3', 'Cross-Disciplinary / Wide Appeal', 'Combines CM/ID/IC aspects; engages different disciplines (nurses, pharmacists) or covers different age groups.'), ('C4', 'Basic + Translational + Clinical Integration', 'Appropriately combines basic, translational, applied, and clinical aspects where applicable.'), ('C5', 'Appropriate Session Format', 'The chosen session format is appropriate to meet the stated session goals.'), ('C6', 'Relevant Collaborators Involved', 'Relevant Study Groups, Affiliated Societies, or other organisations are appropriately involved.'), ('C7', 'Adheres to Session Format Rules', 'Strictly adheres to session format guidelines (number of talks, duration, etc.).'), ('C8', 'Proper Description', 'Proper description provided, highlighting importance, relevance, and educational goals.'), ('C9', 'Gender & Geographic Balance', 'Adequate gender and geographic balance. Underrepresented: young speakers, female speakers, Eastern Europe, low-resource countries.'), ('C10', 'Best Speakers / Not Self-Serving', 'Best available speakers chosen. Proposal is free from self-serving bias.'), ('C11', 'Engages Young Investigators', "Engages young investigators or 'new faces' where applicable.")]

    wb  = Workbook()
    ws  = wb.active
    ws.title = "Proposals"
    thin = Side(style="thin", color="DDDDDD")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    FIELDS = [
        ("ID",               "proposal_id",        7,  "center"),
        ("Session Title",    "session_title",      42, "left"),
        ("Category",         "category",           28, "left"),
        ("Subcategory",      "subcategory",        22, "left"),
        ("Session Type",     "session_type",       18, "left"),
        ("Proposing Entity", "proposing_entities", 24, "left"),
        ("Chairs",           "chairs",             32, "left"),
        ("Reserve Chair",    "reserve_chairs",     22, "left"),
        ("Champion",         "champion",           22, "left"),
        ("# Topics",         "num_topics",          9, "center"),
        ("Topic Titles",     "topic_titles",       55, "left"),
        ("Speakers",         "topic_speakers",     45, "left"),
        ("Description",      "motivation",         55, "left"),
        ("Tags",             "tags",               40, "left"),
        ("Novelty (0-1)",    "score_novelty",      12, "center"),
        ("Trend (-1,+1)",    "score_trend",        12, "center"),
        ("Top Similar",      "sim_summary",        48, "left"),
    ]
    CRIT_START = len(FIELDS) + 1

    def hdr(row, col, val, bg="1F4E79"):
        c = ws.cell(row, col, val)
        c.fill = PatternFill("solid", start_color=bg)
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bdr

    # Row 1: group banners
    ws.row_dimensions[1].height = 18
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(FIELDS))
    c = ws.cell(1, 1, "PROPOSAL DETAILS")
    c.fill = PatternFill("solid", start_color="0F2E4D")
    c.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=CRIT_START,
                   end_row=1, end_column=CRIT_START + len(CRITERIA) - 1)
    c2 = ws.cell(1, CRIT_START,
                 "SCORING CRITERIA  (1=Poor  3=Acceptable  5=Excellent  N/A=Not applicable)")
    c2.fill = PatternFill("solid", start_color="1C4587")
    c2.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    c2.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2: column headers
    ws.row_dimensions[2].height = 42
    for j, (lbl, _, w, _) in enumerate(FIELDS, 1):
        hdr(2, j, lbl)
        ws.column_dimensions[get_column_letter(j)].width = w
    for j, (code, short, _) in enumerate(CRITERIA, CRIT_START):
        c = ws.cell(2, j, f"{code}\n{short}")
        c.fill = PatternFill("solid", start_color="2A5298")
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=8)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bdr
        ws.column_dimensions[get_column_letter(j)].width = 14

    ws.freeze_panes = "A3"

    # Data rows
    for i, p in enumerate(proposals, 3):
        ws.row_dimensions[i].height = 55
        fill = PatternFill("solid", start_color="EEF4FF" if i%2==0 else "F8FAFF")
        sim  = p.get("similar_sessions", [])
        sim_str = "; ".join(
            f"{s['year']} {s['code']}: {s['title'][:40]} ({int(s['similarity']*100)}%)"
            for s in sim[:2]) if sim else ""
        vals = {
            "proposal_id":       p.get("proposal_id",""),
            "session_title":     p.get("session_title",""),
            "category":          p.get("category",""),
            "subcategory":       p.get("subcategory",""),
            "session_type":      p.get("session_type",""),
            "proposing_entities":p.get("proposing_entities",""),
            "chairs":            p.get("chairs",""),
            "reserve_chairs":    p.get("reserve_chairs",""),
            "champion":          p.get("champion",""),
            "num_topics":        p.get("num_topics",""),
            "topic_titles":      p.get("topic_titles",""),
            "topic_speakers":    p.get("topic_speakers",""),
            "motivation":        (p.get("motivation") or "")[:600],
            "tags":              "; ".join(p.get("tags") or []),
            "score_novelty":     p.get("score_novelty",""),
            "score_trend":       p.get("score_trend",""),
            "sim_summary":       sim_str,
        }
        for j, (_, key, _, align) in enumerate(FIELDS, 1):
            c = ws.cell(i, j, vals.get(key,""))
            c.fill = fill
            c.font = Font(name="Arial", size=9)
            c.border = bdr
            c.alignment = Alignment(wrap_text=True, vertical="top",
                                    horizontal="center" if align=="center" else "left")
        crit_fill = PatternFill("solid", start_color="FFFDE7")
        for j in range(CRIT_START, CRIT_START + len(CRITERIA)):
            c = ws.cell(i, j, "")
            c.fill = crit_fill
            c.border = bdr
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(name="Arial", size=11)

    # Sheet 2: Scoring Guide
    ws2 = wb.create_sheet("Scoring Guide")
    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 32
    ws2.column_dimensions["C"].width = 72
    ws2.column_dimensions["D"].width = 22
    t = ws2.cell(1, 1, "ESCMID 2027 EPC — Session Proposal Scoring Criteria")
    t.font = Font(bold=True, color="1F4E79", name="Arial", size=14)
    ws2.merge_cells("A1:D1")
    ws2.row_dimensions[1].height = 24
    sub = ws2.cell(2, 1, "Score each criterion 1 (poor) — 3 (acceptable) — 5 (excellent) — N/A (not applicable)")
    sub.font = Font(italic=True, color="595959", name="Arial", size=10)
    ws2.merge_cells("A2:D2")
    for j, lbl in enumerate(["Code","Criterion","Description","Score Range"], 1):
        c = ws2.cell(4, j, lbl)
        c.fill = PatternFill("solid", start_color="1F4E79")
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr
    ws2.row_dimensions[4].height = 22
    fills2 = [PatternFill("solid", start_color="EEF4FF"),
              PatternFill("solid", start_color="F8FAFF")]
    for i, (code, short, desc) in enumerate(CRITERIA, 5):
        ws2.row_dimensions[i].height = 52
        for j, val in enumerate([code, short, desc, "1/2/3/4/5/N/A"], 1):
            c = ws2.cell(i, j, val)
            c.fill = fills2[i % 2]
            c.font = Font(name="Arial", size=10)
            c.border = bdr
            c.alignment = Alignment(wrap_text=True, vertical="top",
                                    horizontal="center" if j in (1,4) else "left")
    ws2.freeze_panes = "A5"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"  Excel  → {path}")

def build_dashboard(proposals: list, hist_sessions: list, path: Path):
    """Generate the proposal review HTML dashboard."""
    hist_tag_counts = Counter(
        t for s in hist_sessions for t in s.get("tags", [])
    )
    data = {
        "proposals"       : proposals,
        "hist_tag_counts" : dict(hist_tag_counts.most_common(50)),
        "tag_palette"     : TAG_PALETTE,
        "total_hist"      : len(hist_sessions),
        "years"           : sorted(set(s.get("year","") for s in hist_sessions)),
    }
    html_out = _dashboard_html(json.dumps(data, ensure_ascii=True))
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(html_out, encoding="utf-8")
    print(f"  Dashboard → {path}")





def _dashboard_html(data_json: str) -> str:
    """Delegate to build_review_html.generate_html() for clean template handling."""
    import json as _j
    data = _j.loads(data_json)
    try:
        import build_review_html as _brh
        import importlib as _il; _il.reload(_brh)
        return _brh.generate_html(data)
    except ImportError:
        # build_review_html.py not in path — minimal fallback
        n = len(data.get("proposals", []))
        return (
            "<!DOCTYPE html><html><head><meta charset=\"UTF-8\"><title>ESCMID 2027</title></head>"
            f"<body style=\"background:#080e1a;color:#c8d8ec;font-family:sans-serif;padding:40px\">"
            f"<h2>ESCMID 2027 Proposal Review</h2>"
            f"<p>{n} proposals processed. Place <code>build_review_html.py</code> in the same "
            f"directory and rerun to get the full interactive dashboard.</p>"
            f"<script>const D=" + data_json + ";</script></body></html>"
        )

def main():
    ap = argparse.ArgumentParser(description="ESCMID Proposal Reviewer")
    ap.add_argument("--proposals",    default=str(PROPOSALS_DIR),
                    help=f"Folder of proposal HTML files (default: {PROPOSALS_DIR})")
    ap.add_argument("--sessions",     default=str(SESSIONS_JSON),
                    help=f"Path to sessions_raw.json (default: {SESSIONS_JSON})")
    ap.add_argument("--output-dir",   default=str(OUTPUT_DIR))
    ap.add_argument("--ai-review",    action="store_true",
                    help="Use Claude API to rate each proposal on all 11 EPC criteria (~$1-2)")
    ap.add_argument("--skip-tagging", action="store_true",
                    help="Use keyword fallback only (no API)")
    args = ap.parse_args()

    proposals_dir = Path(args.proposals)
    sessions_path = Path(args.sessions)
    out_dir       = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    print("=" * 58)
    print(" ESCMID 2027 Proposal Reviewer")
    print("=" * 58)

    # ── 1. Parse proposals ────────────────────────────────────────────────────
    print(f"\n[1/4] Parsing proposals from {proposals_dir} …")
    if not HAS_PARSER:
        print("  ERROR: parse_proposals_vscode.py not found in same directory")
        sys.exit(1)

    # Only process individual proposal pages, not index/list pages
    html_files = sorted(f for f in proposals_dir.glob("*.html")
                        if "proposal_view" in f.name or "id_" in f.name)
    if not html_files:
        print(f"  No .html files found in {proposals_dir}")
        print("  Save proposal pages as view-source HTML files and place them there.")
        sys.exit(1)

    proposals, errors = [], []
    for path in html_files:
        try:
            inner = _parser.extract_inner_html(path)
            data  = _parser.parse_proposal(inner)
            data["tags"] = []  # will be filled below
            proposals.append(data)
        except Exception as e:
            errors.append((path.name, str(e)))

    print(f"  {len(proposals)} proposals parsed, {len(errors)} errors")

    # ── 2. Load historical sessions ───────────────────────────────────────────
    print(f"\n[2/4] Loading historical sessions from {sessions_path} …")
    if not sessions_path.exists():
        print(f"  Warning: {sessions_path} not found — scores will be zero")
        hist_sessions = []
    else:
        hist_sessions = json.loads(sessions_path.read_text())
        tagged = sum(1 for s in hist_sessions if s.get("tags"))
        print(f"  {len(hist_sessions)} historical sessions ({tagged} with tags)")

    # ── 3. Tag proposals ──────────────────────────────────────────────────────
    print(f"\n[3/4] Tagging proposals …")
    proposals = tag_all_proposals(proposals, use_api=not args.skip_tagging)

    # ── 4. Score proposals ────────────────────────────────────────────────────
    print(f"\n[4/4] Scoring and rating …")
    proposals = score_proposals(proposals, hist_sessions)
    proposals = rate_all_proposals(proposals, use_api=args.ai_review)

    # Save intermediate JSON
    tagged_json = out_dir / "proposals_tagged.json"
    tagged_json.write_text(json.dumps(proposals, indent=2))
    print(f"  JSON  → {tagged_json}")

    # Outputs
    build_excel(proposals, out_dir / "Proposal_Review.xlsx")
    build_dashboard(proposals, hist_sessions, out_dir / "Proposal_Review.html")

    print(f"\nDone — {len(proposals)} proposals reviewed.")
    print(f"Outputs in {out_dir}")


if __name__ == "__main__":
    main()
