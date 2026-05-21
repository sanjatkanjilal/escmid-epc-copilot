"""
ESCMID Proposal HTML Parser — VS Code version
==============================================
Usage:
    python parse_proposals.py                          # uses ./html_files/ -> ./output/
    python parse_proposals.py --input ./my_html_dir   # custom input folder
    python parse_proposals.py --input ./html --output ./results/ratings.xlsx
    python parse_proposals.py --file proposal_6341.html  # single file

Requirements:
    pip install beautifulsoup4 openpyxl
"""

import argparse
import glob
import html
import re
import sys
from pathlib import Path

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Defaults ──────────────────────────────────────────────────────────────────
SCRIPT_DIR     = Path(__file__).parent
DEFAULT_INPUT  = SCRIPT_DIR.parent / "html_files"
DEFAULT_OUTPUT = SCRIPT_DIR.parent / "ESCMID_Session_Rating_Sheet.xlsx"


# ── HTML extraction ───────────────────────────────────────────────────────────

def extract_inner_html(view_source_path: Path) -> str:
    """
    Chrome view-source files wrap the page in a <table> with
    <td class="line-content"> cells.  Reconstruct the original HTML
    by joining those cells and unescaping HTML entities.
    """
    raw = view_source_path.read_text(encoding="utf-8")
    outer = BeautifulSoup(raw, "html.parser")
    lines = [td.get_text() for td in outer.find_all("td", class_="line-content")]
    return html.unescape("\n".join(lines))


# ── Helpers ───────────────────────────────────────────────────────────────────

def clean(text: str) -> str:
    """Collapse whitespace and strip."""
    return re.sub(r"\s+", " ", text or "").strip()


def _person_entry(details_div) -> tuple[str, str]:
    """Return (name_location, email) from a person_details div."""
    texts = list(details_div.stripped_strings)
    return texts[0] if texts else "", texts[1] if len(texts) > 1 else ""


# ── Core parser ───────────────────────────────────────────────────────────────

def parse_proposal(html_content: str) -> dict:
    """
    Parse a single ESCMID proposal detail page.
    Handles the trigger_content div structure (fields are one level deeper
    than proposal_box, inside trigger_content child divs).
    """
    soup = BeautifulSoup(html_content, "html.parser")
    data: dict = {}

    # ── Proposal ID ──────────────────────────────────────────────────────────
    m = (re.search(r"Proposal/Session \((\d+)\)", html_content) or
         re.search(r"rate-proposal-(\d+)", html_content) or
         re.search(r"proposal_view/id/(\d+)", html_content))
    data["proposal_id"] = m.group(1) if m else ""

    # ── Fields via trigger_content (recursive) ────────────────────────────────
    pbox = soup.find("div", class_=re.compile(r"proposal_box"))
    container = (pbox.find("div", class_="trigger_content") or pbox) if pbox else soup

    FIELD_MAP = {
        "Session Title":                 "session_title",
        "Category":                      "category",
        "Subcategory":                   "subcategory",
        "Session Type":                  "session_type",
        "Proposing Entities":            "proposing_entities",
        "Motivation/aim of this session":"motivation",
    }

    for div in container.find_all("div", recursive=True):
        strong = div.find("strong", recursive=False)
        if not strong:
            continue
        label = clean(strong.get_text())
        if label not in FIELD_MAP:
            continue
        p = div.find("p")
        if p:
            val = clean(p.get_text())
            if val:
                data[FIELD_MAP[label]] = val

    # ── Chairs / reserve chairs / champion ────────────────────────────────────
    chairs, reserve_chairs = [], []
    champion = ""
    for pd in soup.find_all("div", class_="person_details"):
        texts = [t.strip() for t in pd.stripped_strings if t.strip()]
        if not texts:
            continue
        role  = texts[0].lower()
        name  = texts[1] if len(texts) > 1 else ""
        email = texts[2] if len(texts) > 2 else ""
        # Email sometimes has trailing junk after a comma
        email = email.split(",")[0].strip()
        entry = f"{name} | {email}".strip(" |")
        if "reserve" in role:
            reserve_chairs.append(entry)
        elif "champion" in role:
            champion = entry
        elif "chair" in role:
            chairs.append(entry)

    data["chairs"]         = " ; ".join(chairs)
    data["reserve_chairs"] = " ; ".join(reserve_chairs)
    data["champion"]       = champion

    # ── Topics & speakers ─────────────────────────────────────────────────────
    topics = []
    for topic_div in soup.find_all("div", class_="topic"):
        # Title: left half div (absent in keynotes — no title, just speaker)
        title_div = topic_div.find(
            "div", class_=re.compile(r"\bleft\b.*\bhalf\b|\bhalf\b.*\bleft\b")
        )
        title = clean(title_div.get_text()) if title_div else ""

        speakers = []
        spk_block = topic_div.find("div", class_="topic_persons_main")
        if spk_block:
            for pd in spk_block.find_all("div", class_="person_details"):
                spk_texts = [t.strip() for t in pd.stripped_strings if t.strip()]
                if not spk_texts:
                    continue
                # Speakers have no role prefix — first text IS the name
                # Guard against role words accidentally present
                name = (spk_texts[0]
                        if not re.match(r"^(chair|reserve|champion|speaker)",
                                        spk_texts[0].lower())
                        else (spk_texts[1] if len(spk_texts) > 1 else ""))
                if name:
                    speakers.append(name)

        topics.append({"title": title, "speakers": " ; ".join(speakers)})

    data["num_topics"]     = len(topics)
    data["topic_titles"]   = " | ".join(t["title"]   for t in topics)
    data["topic_speakers"] = " | ".join(t["speakers"] for t in topics)

    return data

# ── Excel builder ─────────────────────────────────────────────────────────────

COLUMNS = [
    ("ID",                    6,  "center"),
    ("Session Title",         40, "left"),
    ("Category",              28, "left"),
    ("Session Type",          20, "left"),
    ("Proposing Entity",      22, "left"),
    ("Chairs",                30, "left"),
    ("Reserve Chair",         22, "left"),
    ("Champion",              22, "left"),
    ("# Topics",              10, "center"),
    ("Topic Titles",          50, "left"),
    ("Speakers",              45, "left"),
    ("Motivation (Summary)",  55, "left"),
    ("⭐ Rating (1–5)",       14, "center"),
    ("✅ Recommend?",         14, "center"),
    ("💬 Comments",           40, "left"),
]

RATING_COLS = {13, 14, 15}


def build_excel(proposals: list[dict], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Session Ratings"

    # Styles
    hdr_font    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill    = PatternFill("solid", start_color="1F4E79")
    rating_fill = PatternFill("solid", start_color="FFF2CC")
    rating_hdr  = PatternFill("solid", start_color="375623")
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    thin        = Side(style="thin", color="BFBFBF")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header row
    ws.row_dimensions[1].height = 32
    for col_idx, (header, width, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = hdr_font
        cell.fill      = rating_hdr if col_idx in RATING_COLS else hdr_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for row_idx, p in enumerate(proposals, start=2):
        ws.row_dimensions[row_idx].height = 60
        row_data = [
            p.get("proposal_id",       ""),
            p.get("session_title",     ""),
            p.get("category",          ""),
            p.get("session_type",      ""),
            p.get("proposing_entities",""),
            p.get("chairs",            ""),
            p.get("reserve_chairs",    ""),
            p.get("champion",          ""),
            p.get("num_topics",        ""),
            p.get("topic_titles",      ""),
            p.get("topic_speakers",    ""),
            (p.get("motivation") or "")[:500],
            "",  # Rating
            "",  # Recommend
            "",  # Comments
        ]
        for col_idx, value in enumerate(row_data, start=1):
            _, _, align = COLUMNS[col_idx - 1]
            cell            = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border     = border
            cell.alignment  = center if align == "center" else left_wrap
            cell.font       = Font(name="Arial", size=9)
            if col_idx in RATING_COLS:
                cell.fill = rating_fill

    ws.freeze_panes = "A2"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "ESCMID Session Rating Summary"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws2["A2"] = f"Total sessions: {len(proposals)}"
    ws2["A2"].font = Font(name="Arial", size=11)
    ws2["A3"].font = Font(name="Arial", bold=True, size=11)
    ws2["A3"] = "Average Rating:"
    ws2["B3"] = f"=IFERROR(AVERAGE('Session Ratings'!M2:M{len(proposals)+1}),\"No ratings yet\")"
    ws2["B3"].font = Font(name="Arial", size=11)

    wb.save(output_path)
    print(f"✅ Saved → {output_path}")


# ── CLI entry point ───────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Parse ESCMID proposal HTML files into a rating Excel sheet."
    )
    group = parser.add_mutually_exclusive_group()
    group.add_argument(
        "--input", "-i",
        type=Path,
        default=DEFAULT_INPUT,
        help=f"Folder containing HTML files (default: {DEFAULT_INPUT})",
    )
    group.add_argument(
        "--file", "-f",
        type=Path,
        default=None,
        help="Parse a single HTML file instead of a folder",
    )
    parser.add_argument(
        "--output", "-o",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Output Excel path (default: {DEFAULT_OUTPUT})",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    # Collect files
    if args.file:
        if not args.file.exists():
            sys.exit(f"❌ File not found: {args.file}")
        html_files = [args.file]
    else:
        if not args.input.exists():
            sys.exit(
                f"❌ Input folder not found: {args.input}\n"
                f"   Create it and drop your HTML files in, or use --input <path>"
            )
        html_files = sorted(args.input.glob("*.html"))
        if not html_files:
            sys.exit(f"❌ No .html files found in {args.input}")

    print(f"📂 Found {len(html_files)} HTML file(s)")

    proposals = []
    errors    = []
    for path in html_files:
        print(f"   Parsing: {path.name} … ", end="", flush=True)
        try:
            inner = extract_inner_html(path)
            data  = parse_proposal(inner)
            proposals.append(data)
            print(f"ID={data.get('proposal_id')} | {data.get('session_title','')[:55]}")
        except Exception as exc:
            print(f"ERROR — {exc}")
            errors.append((path.name, str(exc)))

    if not proposals:
        sys.exit("❌ No proposals parsed successfully.")

    proposals.sort(key=lambda x: x.get("proposal_id", ""))
    build_excel(proposals, args.output)
    print(f"\n🎉 Done — {len(proposals)} proposals, {len(errors)} error(s).")
    if errors:
        print("\nFiles with errors:")
        for fname, err in errors:
            print(f"  {fname}: {err}")


if __name__ == "__main__":
    main()
