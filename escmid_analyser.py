#!/usr/bin/env python3
"""
ESCMID EPC Copilot
==================
Extracts all sessions from ESCMID annual programme PDFs,
classifies them using the official ESCMID taxonomy via the
Anthropic API, and produces:

  1. An Excel workbook  — one sheet per ESCMID category (12) + overview
  2. An interactive HTML dashboard — trends, heatmap, network, explore, gaps

Quick start
-----------
1. Install dependencies:
       pip install -r requirements.txt
       # pdfplumber is required for PDF extraction (replaces pdftotext)
       # macOS: brew install poppler   (for pdftotext)
       # Ubuntu: sudo apt install poppler-utils
       # Windows: https://github.com/oschwartz10612/poppler-windows

2. Set your Anthropic API key — either in a .env file:
       echo 'ANTHROPIC_API_KEY=sk-ant-...' > .env
   or in your shell:
       export ANTHROPIC_API_KEY="sk-ant-..."

3. Place PDF files in  data/programmes/  (configure PROGRAMME_FILES below)

4. Run:
       python escmid_analyser.py
       python escmid_analyser.py --years 2025,2026          # specific years
       python escmid_analyser.py --skip-tagging              # keyword fallback only
       python escmid_analyser.py --skip-excel                # dashboard only

Outputs land in  data/output/
  ESCMID_Programmes.xlsx
  ESCMID_Dashboard.html
  sessions_raw.json         (intermediate — safe to delete)
  tagging_cache.json        (API responses cached here — keep to avoid re-billing)

Improving the script
--------------------
- Add / remove PROGRAMME_FILES entries to cover more years
- Tune KEYWORD_RULES for better fallback accuracy
- Modify INCLUDE_TYPES to control which session formats are extracted
- The API tagging prompt is in build_tagging_prompt() — edit it to improve accuracy
- The HTML dashboard template is in generate_dashboard_html()

"""

# ── Standard library ──────────────────────────────────────────────────────────
import os, re, json, time, subprocess, sys, argparse
from pathlib import Path
from collections import defaultdict, Counter

# ── Load .env file if present (pip install python-dotenv) ─────────────────────
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # .env loading is optional; key can also be set in the shell directly

# ── Optional third-party ──────────────────────────────────────────────────────
try:
    import anthropic as _anthropic
    HAS_ANTHROPIC = True
except ImportError:
    HAS_ANTHROPIC = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from tqdm import tqdm as _tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False
    _tqdm = lambda x, **_: x


# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION — edit this section for your setup
# ══════════════════════════════════════════════════════════════════════════════

BASE_DIR      = Path(__file__).parent
PROGRAMMES_DIR = BASE_DIR / "data" / "programmes"
OUTPUT_DIR    = BASE_DIR / "data" / "output"
CACHE_FILE    = BASE_DIR / "data" / "tagging_cache.json"

# Map year → PDF filename (relative to PROGRAMMES_DIR)
PROGRAMME_FILES = {
    "2021": "FinalProgramme_2021.pdf",
    "2022": "FinalProgramme_2022_Lisbon.pdf",
    "2023": "FinalProgramme_2023_Copenhagen.pdf",
    "2024": "FinalProgramme_2024_Barcelona.pdf",
    "2025": "FinalProgramme_2025_Vienna.pdf",
    "2026": "FinalProgramme_2026_Munich.pdf",
}

# Session type codes to include (prefix of session codes like OS001, SY010…)
# 2021 uses S## format; 2022+ uses OS/SY/EW/ME/KN/LB/EF/PM/CS/IS
INCLUDE_TYPES_MODERN = {"OS","SY","EW","ME","KN","LB","EF","PM","CS","IS"}
INCLUDE_TYPES_2021   = {"S"}    # 2021 uses S## codes

# Session code prefix → canonical session type
# These prefixes are consistent across all ECCMID/ESCMID years.
# Used as a reliable fallback when the type cannot be inferred from context text.
PREFIX_TO_TYPE = {
    "OS": "Oral Session",    "SY": "Symposium",
    "EW": "Educational",     "ME": "Meet-the-Expert",
    "KN": "Keynote",         "LB": "Late-Breaking",
    "EF": "ePoster Flash",   "PM": "Pipeline",
    "CS": "Case Session",    "IS": "IS / Integrated",
}

# Session types shown in the clustering network (keep manageable)
NETWORK_TYPE_NAMES = {"Oral Session","Symposium","Keynote","Late-Breaking","Pipeline","Meet-the-Expert","Case Session"}

# Anthropic API settings
ANTHROPIC_MODEL   = "claude-sonnet-4-20250514"
API_DELAY         = 0.4   # seconds between calls
MAX_RETRIES       = 3
SAVE_CACHE_EVERY  = 25    # write cache every N sessions


# ══════════════════════════════════════════════════════════════════════════════
# ESCMID OFFICIAL TAXONOMY
# Source: https://www.escmid.org/congress-events/escmid-global/proposal-entry/
#         escmid-global-categories-and-subcategories/
# ══════════════════════════════════════════════════════════════════════════════

CATEGORIES = {
    1:  {"name":"Viral infection & disease",
         "short":"Viral disease","color":"#4fc3f7","subcats":{
         "1a":"HIV/AIDS","1b":"Viral hepatitis","1c":"Influenza & respiratory viruses",
         "1d":"Herpesviruses","1e":"Emerging/re-emerging viral diseases","1f":"Diagnostic virology (other)",
         "1g":"Viral epidemiology – general","1h":"Antiviral drugs (other)",
         "1i":"Fundamental & applied virology","1j":"COVID-19"}},
    2:  {"name":"Bacterial infection & disease",
         "short":"Bacterial disease","color":"#fbbf24","subcats":{
         "2a":"TB & other mycobacterial infections","2b":"Severe sepsis, bacteraemia & endocarditis",
         "2c":"Community-acquired respiratory infections","2d":"Community-acquired abdominal/GI infections",
         "2e":"Community-acquired UTI & genital tract","2f":"Community-acquired SSTI, bone & joint",
         "2g":"Community-acquired CNS & invasive infections","2h":"Zoonotic bacterial infections",
         "2i":"Other intracellular or rare bacteria"}},
    3:  {"name":"Bacterial susceptibility & resistance",
         "short":"AMR","color":"#f87171","subcats":{
         "3a":"Resistance surveillance: community","3b":"Resistance surveillance: healthcare",
         "3c":"Susceptibility testing methods","3d":"Resistance mechanisms",
         "3e":"Resistance detection/prediction","3f":"Clinical outcome of resistant infections",
         "3g":"Spread of resistance (ecology, One Health)","3h":"Policy aspects of AMR"}},
    4:  {"name":"Diagnostic microbiology",
         "short":"Diagnostics","color":"#a78bfa","subcats":{
         "4a":"Diagnostic bacteriology","4b":"Laboratory management",
         "4c":"MALDI-TOF & proteomic methods","4d":"Molecular diagnostics (incl POCT)",
         "4e":"Strain typing & surveillance","4f":"Whole genome sequencing",
         "4g":"Microbiome studies","4h":"Clinical metagenomics",
         "4i":"Bioinformatics tools & pipelines","4j":"Artificial intelligence & digital health",
         "4k":"Other novel diagnostic technologies"}},
    5:  {"name":"New antibacterial agents, PK/PD & stewardship",
         "short":"Antibacterials & AMS","color":"#34d399","subcats":{
         "5a":"Drug discovery & new compounds","5b":"Pharmacokinetics/pharmacodynamics",
         "5c":"New/repurposed agents: clinical studies","5d":"Antimicrobial stewardship & prescribing",
         "5e":"Safety, hypersensitivity & adverse effects","5f":"Pharmacoepidemiology/pharmacoeconomics"}},
    6:  {"name":"Fungal infection & disease",
         "short":"Mycology","color":"#fb923c","subcats":{
         "6a":"Fundamental mycology","6b":"Fungal disease epidemiology",
         "6c":"Diagnostic mycology","6d":"Antifungal susceptibility & resistance",
         "6e":"Antifungal drugs & treatment"}},
    7:  {"name":"Parasitic diseases, travel medicine & migrant health",
         "short":"Parasitology & Travel","color":"#6ee7b7","subcats":{
         "7a":"Fundamental parasitology","7b":"Parasitic disease epidemiology",
         "7c":"Diagnostic parasitology","7d":"Antiparasitic drugs & treatment",
         "7e":"Antiparasitic susceptibility & resistance","7f":"Travel medicine & migrant health"}},
    8:  {"name":"Healthcare-associated infections & IPC",
         "short":"HAI & IPC","color":"#38bdf8","subcats":{
         "8a":"Intravascular catheter-related infections","8b":"Foreign-body & implant infections",
         "8c":"Surgical site infections","8d":"Healthcare-associated pneumonia (VAP)",
         "8e":"Hospital epidemiology, transmission & surveillance",
         "8f":"Other HAIs (CDI, outbreaks)","8g":"Infection control interventions & trials",
         "8h":"Disinfection & sterilisation","8i":"Healthcare workers & IPC"}},
    9:  {"name":"Fundamental microbiology, pathogenesis & immunity",
         "short":"Basic science","color":"#e879f9","subcats":{
         "9a":"Microbial pathogenesis & virulence","9b":"Host-pathogen interaction",
         "9c":"Pre-clinical biofilm studies","9d":"Experimental & cellular microbiology",
         "9e":"Fundamental science using Omics","9f":"Immune response to infection"}},
    10: {"name":"Immune compromise & transplant ID",
         "short":"Immunocompromised","color":"#f472b6","subcats":{
         "10a":"Host genetics & primary immunodeficiency","10b":"Solid organ transplantation",
         "10c":"Haematopoietic stem cell transplantation","10d":"Cell-based therapies",
         "10e":"Cancer treatment & neutropaenia","10f":"Other immunosuppression"}},
    11: {"name":"Public health & vaccines",
         "short":"Public health","color":"#86efac","subcats":{
         "11a":"General vaccinology","11b":"Antiviral vaccines","11c":"Antibacterial vaccines",
         "11d":"Other preventive modalities","11e":"Food, water & environmental health",
         "11f":"Veterinary microbiology & One Health","11g":"Global health & health security",
         "11h":"Infections in low-resource settings"}},
    12: {"name":"Professional & educational affairs",
         "short":"Professional","color":"#94a3b8","subcats":{
         "12a":"Professional affairs & career development","12b":"Publishing, ethics & academic affairs",
         "12c":"Medical education for CM/ID","12d":"Diversity & equality",
         "12e":"Advocacy & role of patients"}},
}

# Flat lookup: subcat code → (cat_num, cat_name, subcat_name)
SUBCAT_LOOKUP = {
    sc: (cn, cat["name"], sn)
    for cn, cat in CATEGORIES.items()
    for sc, sn in cat["subcats"].items()
}

# Prompt string: all subcategories for the API
ALL_SUBCATS_STR = "\n".join(
    f"  {sc}: {sn}  [{CATEGORIES[cn]['name']}]"
    for cn, cat in CATEGORIES.items()
    for sc, sn in cat["subcats"].items()
)




# ── Legacy pdftotext extraction helpers (used for 2021 and fallback) ──────────
_Y2021_CODE    = re.compile(r'\b(S\d{2,3})\b')
MODERN_CODE_RE = re.compile(
    r'\b((?:OS|SY|EW|ME|LB|EF|KN|PM|CS|IS)\d{2,4})\b'
)
INCLUDE_TYPES_MODERN = {"OS","SY","EW","ME","KN","LB","EF","PM","CS","IS"}

def classify_session_type(text: str) -> str:
    lc = text.lower()
    for label, kws in [
        ("Oral Session",    ["oral session","oral case","mini oral"]),
        ("Symposium",       ["symposium"]),
        ("Educational",     ["educational","workshop"]),
        ("Meet-the-Expert", ["meet-the-expert","meet the expert"]),
        ("Keynote",         ["keynote"]),
        ("Late-Breaking",   ["late-breaking","late breaking"]),
        ("ePoster Flash",   ["eposter flash","poster flash","eposter review"]),
        ("Pipeline",        ["pipeline"]),
    ]:
        if any(k in lc for k in kws):
            return label
    return "Other"


def _stitch_talk(parts: list) -> str:
    """Join word-wrapped fragments, removing mid-word hyphens."""
    result = ""
    for p in (p.strip() for p in parts if p.strip()):
        if result.endswith("-"):
            result = result[:-1] + p
        elif result:
            result += " " + p
        else:
            result = p
    return result.strip()


def pdf_to_text(path: Path) -> str:
    """Convert PDF to text using pdftotext (used for 2021 fallback)."""
    result = subprocess.run(
        ["pdftotext", "-layout", str(path), "-"],
        capture_output=True, text=True, timeout=120
    )
    if result.returncode != 0:
        raise RuntimeError(f"pdftotext failed: {result.stderr[:200]}")
    return result.stdout


def extract_sessions_modern(text: str, year: str) -> list:
    """pdftotext-based extractor — fallback when pdfplumber unavailable."""
    lines = text.split("\n")
    sessions, seen = [], set()
    for i, line in enumerate(lines):
        for m in MODERN_CODE_RE.finditer(line):
            code = m.group(1)
            if code[:2] not in INCLUDE_TYPES_MODERN or code in seen:
                continue
            ctx = lines[i:i+8]
            title, stype = "", "Other"
            for cl in ctx[1:]:
                c = cl.strip()
                if (len(c) > 10 and not MODERN_CODE_RE.match(c)
                        and not re.match(r'^\d{2}:\d{2}', c)):
                    title = re.split(r'\s{3,}', c)[0].strip()
                    stype = classify_session_type(" ".join(ctx))
                    break
            if title:
                seen.add(code)
                sessions.append({"code": code, "year": year,
                                  "type": stype, "title": title, "talks": []})
    return sessions


def extract_sessions_2021(text: str) -> list:
    """Extract sessions from 2021 online ECCMID PDF (S## code format)."""
    lines = text.split("\n")
    sessions, seen = [], set()
    for i, line in enumerate(lines):
        for m in _Y2021_CODE.finditer(line):
            code = m.group(1)
            if code in seen or not re.search(r'\d{2}:\d{2}', line):
                continue
            col = max(0, m.start() - 2)
            ctx_parts = []
            for j in range(i, min(i + 8, len(lines))):
                seg = lines[j][col:col + 28].strip()
                if seg:
                    ctx_parts.append(seg)
            title, stype, frags = "", "Other", []
            for part in ctx_parts[1:]:
                if re.match(r'^\d{2}:\d{2}', part):
                    continue
                if _Y2021_CODE.match(part):
                    break
                if any(k in part.lower() for k in ["oral","session","symposium",
                       "workshop","keynote","eposter","poster review"]):
                    stype = classify_session_type(part)
                    continue
                if len(part) > 6:
                    frags.append(part)
                    if len(frags) >= 3:
                        break
            title = _stitch_talk(frags)[:200]
            if title and not re.match(r'^[A-Z]\. [A-Z]', title):
                seen.add(code)
                sessions.append({"code": code, "year": "2021",
                                  "type": stype, "title": title, "talks": []})
    return sessions


# ══════════════════════════════════════════════════════════════════════════════
# PDF EXTRACTION — pdfplumber engine (2022-2026)
# Uses word-level bounding boxes for clean column separation.
# ══════════════════════════════════════════════════════════════════════════════

try:
    import pdfplumber as _pdfplumber
    HAS_PLUMBER = True
except ImportError:
    HAS_PLUMBER = False
    print("Warning: pdfplumber not installed. Run: pip install pdfplumber")

# ── Regex patterns for state-machine parser ───────────────────────────────────
_SESS_HDR = re.compile(
    r'^((?:OS|SY|EW|ME|LB|EF|KN|PM|CS|IS)\d{2,4})\s+'
    r'(\d{1,2}:\d{2})\s*[-\u2013]\s*(\d{1,2}:\d{2})'
    r'(?:\s+(.+))?$'
)
_TALK_HDR = re.compile(r'^([A-Z]\d{4})\s+(\d{1,2}:\d{2})\s+(.+)')
_TYPE_RE  = re.compile(
    r'(oral\s+(?:case\s+)?session|symposium|educational(?:\s+session)?'
    r'|workshop|meet.the.expert|keynote|late.breaking|eposter\s+flash'
    r'|pipeline|case\s+session|integrated\s+symposium)',
    re.I
)
_CHAIR_RE = re.compile(r'^Chairs?\s*(.*)', re.I)
_CO_ORG   = re.compile(r'^Co-organised', re.I)
_SPEAKER  = re.compile(r'\*|[A-Z]\.\s+[A-Z]|\([A-Z][a-z]+,')

_PLUMB_TYPE = [
    ('oral',       'Oral Session'),   ('symposium',  'Symposium'),
    ('educational','Educational'),    ('workshop',   'Educational'),
    ('meet',       'Meet-the-Expert'),('keynote',    'Keynote'),
    ('late',       'Late-Breaking'),  ('eposter',    'ePoster Flash'),
    ('pipeline',   'Pipeline'),       ('case',       'Case Session'),
    ('integrated', 'IS / Integrated'),
]

def _plumb_type(line: str) -> str:
    m = _TYPE_RE.search(line.lower())
    if not m: return 'Other'
    raw = m.group(1).lower()
    for prefix, label in _PLUMB_TYPE:
        if raw.startswith(prefix): return label
    return 'Other'


def _finalise_talk(talk: dict):
    if 'title_parts' in talk:
        t = ' '.join(talk.pop('title_parts'))
        t = re.sub(r'(?<=\w)- ', '', t)
        talk['title'] = re.sub(r'\s+', ' ', t).strip()
    if 'speaker_parts' in talk:
        talk['speakers'] = '; '.join(
            s.rstrip(',') for s in talk.pop('speaker_parts') if s
        )


def _parse_column(text: str, year: str) -> list:
    """State-machine parser for a single clean column text stream."""
    sessions: list = []
    lines = [l.rstrip() for l in text.split('\n')]
    cur: dict = None

    def _save():
        if not cur: return
        t = ' '.join(cur.pop('title_parts', []))
        t = re.sub(r'(?<=\w)- ', '', t)
        cur['title']  = re.sub(r'\s+', ' ', t).strip()
        cur['chairs'] = '; '.join(cur.pop('chair_parts', [])).strip('; ')
        for talk in cur.get('talks', []):
            _finalise_talk(talk)
        sessions.append(dict(cur))

    for raw in lines:
        line = raw.strip()
        if not line: continue

        m = _SESS_HDR.match(line)
        if m:
            _save(); cur = None
            cur = {
                'code':        m.group(1),      'year':  year,
                'time_hall':   f"{m.group(2)} \u2013 {m.group(3)}"
                               + (f"  {m.group(4)}" if m.group(4) else ''),
                'type':        'Other',          'title_parts': [],
                'chair_parts': [],               'talks': [],
                '_state':      'type',
            }
            continue

        if cur is None: continue
        state = cur['_state']

        tm = _TALK_HDR.match(line)
        if tm and state in ('title','chairs','talks','talk_title','talk_spk'):
            cur['_state'] = 'talk_title'
            cur['talks'].append({
                'code': tm.group(1), 'time': tm.group(2),
                'title_parts': [tm.group(3)], 'speaker_parts': [],
            })
            continue

        if _CO_ORG.match(line):
            cur['_state'] = 'footer'
            continue

        if state == 'type':
            if _TYPE_RE.search(line):
                cur['type'] = _plumb_type(line)
                cur['_state'] = 'title'

        elif state == 'title':
            cm = _CHAIR_RE.match(line)
            if cm:
                cur['_state'] = 'chairs'
                if cm.group(1): cur['chair_parts'].append(cm.group(1))
            else:
                cur['title_parts'].append(line)

        elif state == 'chairs':
            if not re.match(r'^\(', line):
                cur['chair_parts'].append(line)

        elif state == 'talk_title':
            talk = cur['talks'][-1]
            if _SPEAKER.search(line):
                cur['_state'] = 'talk_spk'
                talk['speaker_parts'].append(line)
            else:
                talk['title_parts'].append(line)

        elif state == 'talk_spk':
            cur['talks'][-1]['speaker_parts'].append(line)

    _save()
    return sessions


def extract_sessions_plumber(path: Path, year: str) -> list:
    """Extract sessions with talks and chairs using pdfplumber column separation."""
    from collections import defaultdict as _dd

    def _w2s(words):
        words = sorted(words, key=lambda w: (round(w['top']/4)*4, w['x0']))
        lines = _dd(list)
        for w in words:
            lines[round(w['top']/4)*4].append(w['text'])
        return '\n'.join(' '.join(lines[k]) for k in sorted(lines))

    col_l, col_r = [], []
    with _pdfplumber.open(str(path)) as pdf:
        for page in pdf.pages:
            words = page.extract_words(x_tolerance=3, y_tolerance=3)
            if not words: continue
            mid = page.width / 2
            col_l.append(_w2s([w for w in words if w['x0'] < mid - 20]))
            col_r.append(_w2s([w for w in words if w['x0'] >= mid - 20]))

    sessions, seen = [], set()
    for text in ['\n'.join(col_l), '\n'.join(col_r)]:
        for s in _parse_column(text, year):
            if s['code'] not in seen:
                seen.add(s['code'])
                s.pop('_state', None)
                sessions.append(s)
    return sessions


def extract_talks_from_text(text: str) -> dict:
    """
    Scan pdftotext output for talk entries (used for 2021 fallback).
    Returns {session_code: [{"code","time","title","speakers"}]}
    """
    lines = text.split("\n")
    _TALK_A = re.compile(r'(?<!\w)([A-Z]\d{4})\s+(\d{1,2}:\d{2})\s+(\S.{4,})')
    _TALK_B = re.compile(r'(?<!\w)([A-Z]\d{4})\t{2,}\s*(\S.{4,})')
    _SPK    = re.compile(r'\*|[A-Z]\.\s+[A-Z]|\([A-Z][a-z]+,')

    sess_pos = {}
    for i, line in enumerate(lines):
        codes = MODERN_CODE_RE.findall(line)
        if len(codes) == 1 and codes[0] not in sess_pos:
            sess_pos[codes[0]] = i
    sorted_sess = sorted(sess_pos.items(), key=lambda x: x[1])

    raw = []
    for i, line in enumerate(lines):
        m = _TALK_A.search(line)
        if m:
            raw.append((i, m.group(1), m.group(2), m.group(3), m.start()))
            continue
        m = _TALK_B.search(line)
        if m:
            raw.append((i, m.group(1), "", m.group(2), m.start()))

    parsed = []
    for idx, (li, code, time, title_start, col) in enumerate(raw):
        title_p, spk_p, seen_spk = [title_start.strip()], [], False
        stop = min(raw[idx+1][0] if idx+1 < len(raw) else li+12, li+12)
        for j in range(li+1, stop):
            nxt = lines[j]
            seg = nxt[col:col+50].strip() if len(nxt) > col else ""
            if not seg:
                if seen_spk: break
                continue
            if MODERN_CODE_RE.search(seg): break
            if _SPK.search(seg) or seen_spk:
                seen_spk = True; spk_p.append(seg)
            else:
                title_p.append(seg)
        parsed.append({"li": li, "code": code, "time": time,
                        "title": _stitch_talk(title_p)[:220],
                        "speakers": "; ".join(s.rstrip(",") for s in spk_p if s)[:200]})

    from collections import defaultdict
    by_sess = defaultdict(list)
    for t in parsed:
        best = None
        for sc, sl in sorted_sess:
            if sl <= t["li"]: best = sc
            else: break
        if best:
            by_sess[best].append({k: v for k, v in t.items() if k != "li"})
    return dict(by_sess)


def extract_all(pdf_map: dict, include_talks: bool = False) -> list:
    """
    Extract sessions from all PDFs.
    2022-2026: pdfplumber (clean columns, talks always included).
    2021:      pdftotext fallback.
    """
    all_sessions = []
    for year, path in sorted(pdf_map.items()):
        if not path.exists():
            print(f"  [SKIP] {year}: file not found ({path})")
            continue
        print(f"  Extracting {year} \u2026", end=" ", flush=True)
        try:
            if year == "2021" or not HAS_PLUMBER:
                text = pdf_to_text(path)
                if year == "2021":
                    sessions = extract_sessions_2021(text)
                else:
                    sessions = extract_sessions_modern(text, year)
                if include_talks:
                    tmap = extract_talks_from_text(text)
                    for s in sessions:
                        s.setdefault("talks", tmap.get(s["code"], []))
                n = sum(len(s.get("talks",[])) for s in sessions)
                print(f"{len(sessions)} sessions" +
                      (f", {n} talks" if include_talks else ""))
            else:
                sessions = extract_sessions_plumber(path, year)
                n = sum(len(s.get("talks",[])) for s in sessions)
                print(f"{len(sessions)} sessions, {n} talks")
            all_sessions.extend(sessions)
        except Exception as exc:
            import traceback
            print(f"ERROR \u2014 {exc}")
            traceback.print_exc()
    return all_sessions



# ══════════════════════════════════════════════════════════════════════════════
# KEYWORD FALLBACK TAGGER
# (used when --skip-tagging or ANTHROPIC_API_KEY not set)
# ══════════════════════════════════════════════════════════════════════════════

# Each rule: (primary_cat, primary_subcat, [keywords…])
# Rules are evaluated in order; first match wins.
KEYWORD_RULES = [
    # ── Category 4j: AI / digital health (check first — cross-cutting) ──────
    (4, "4j", ["artificial intelligence","machine learning","deep learning",
               "neural network","large language model","chatgpt","llm",
               "natural language processing","nlp","digital health"," ai-",
               "ai tool","ai model","generative ai","computer vision"]),
    # ── Category 1: Viral ────────────────────────────────────────────────────
    (1, "1j", ["covid","sars-cov","covid-19","coronavirus"]),
    (1, "1a", ["hiv","aids","antiretroviral","art therapy","pre-exposure prophylaxis"]),
    (1, "1b", ["hepatitis b","hepatitis c","hepatitis e","hbv","hcv","hev","viral hepatitis"]),
    (1, "1c", ["influenza","rsv","respiratory syncytial","respiratory virus","flu vaccine",
               "metapneumovirus","parainfluenza","rhinovirus"]),
    (1, "1d", ["herpes","cmv","cytomegalovirus","ebv","epstein-barr","vzv","hsv",
               "varicella zoster","kaposi"]),
    (1, "1e", ["mpox","monkeypox","ebola","marburg","dengue","zika","west nile",
               "emerging virus","hantavirus","rift valley","nipah"]),
    (1, "1i", ["phage","bacteriophage","phage therapy"]),
    # ── Category 2: Bacterial disease ────────────────────────────────────────
    (2, "2a", ["tuberculosis","mycobacteri","mycobacteria"," tb ","tb treatment",
               "latent tb","non-tuberculous mycobacteria","ntm"]),
    (2, "2b", ["sepsis","bacteraemia","bacteremia","bloodstream infection",
               "endocarditis","infective endocarditis","bacteraemic"]),
    (2, "2c", ["community-acquired pneumonia","cap ","lower respiratory","bronchitis",
               "community respiratory"]),
    (2, "2d", ["intraabdominal","peritonitis","abdominal sepsis","gastrointestinal infection",
               "cdiff","c. difficile","clostridioides","clostridium difficile"]),
    (2, "2e", ["urinary tract infection","uti ","cystitis","pyelonephritis","urinary infection"]),
    (2, "2f", ["skin and soft tissue","cellulitis","necrotising fasciitis","osteomyelitis",
               "septic arthritis","bone infection","joint infection"]),
    (2, "2g", ["meningitis","encephalitis","brain abscess","cns infection","neurological infection"]),
    (2, "2h", ["zoonotic","lyme disease","brucella","leptospira","rickettsial"]),
    # ── Category 3: AMR ───────────────────────────────────────────────────────
    (3, "3a", ["resistance surveillance","amr surveillance","epidemiology of resistance",
               "prevalence of resistance"]),
    (3, "3b", ["healthcare-associated resistance","hospital-acquired resistance","nosocomial amr"]),
    (3, "3c", ["susceptibility testing","mic determination","breakpoint","eucast","clsi",
               "disk diffusion","e-test","broth microdilution"]),
    (3, "3d", ["resistance mechanism","beta-lactamase","carbapenemase","kpc","ndm",
               "oxa-","mcr-","mobile genetic","plasmid transfer"]),
    (3, "3e", ["resistance prediction","resistance detection","rapid amr","resistome"]),
    (3, "3f", ["outcome of resistant","clinical outcome of resistant","mdr infection outcome"]),
    (3, "3g", ["spread of resistance","one health amr","amr ecology","amr reservoir"]),
    (3, "3h", ["amr policy","antimicrobial policy","amr economic","amr governance"]),
    # ── Category 4: Diagnostics ───────────────────────────────────────────────
    (4, "4a", ["blood culture","diagnostic bacteriology","culture-based diagnosis"]),
    (4, "4b", ["laboratory management","lab automation","laboratory quality","laboratory data"]),
    (4, "4c", ["maldi-tof","mass spectrometry","proteomics"]),
    (4, "4d", ["molecular diagnostic","pcr","point-of-care","poct","syndromic testing",
               "multiplex pcr","rapid molecular"]),
    (4, "4e", ["strain typing","genomic epidemiology","molecular typing","wgmlst","cgmlst"]),
    (4, "4f", ["whole genome sequencing","wgs","next-generation sequencing","ngs",
               "nanopore","sequencing-based"]),
    (4, "4g", ["microbiome","gut microbiota","microbiota","dysbiosis","16s rrna"]),
    (4, "4h", ["metagenomics","metagenomic","clinical metagenomics","shotgun sequencing"]),
    (4, "4i", ["bioinformatics","pipeline","software tool","computational"]),
    (4, "4k", ["novel diagnostic","lateral flow","immunochromatography","biosensor"]),
    # ── Category 5: Antibacterials & AMS ──────────────────────────────────────
    (5, "5a", ["drug discovery","new antibiotic","new compound","novel antibiotic",
               "drug design","antibiotic pipeline","antimicrobial candidate"]),
    (5, "5b", ["pharmacokinetics","pharmacodynamics","pk/pd","therapeutic drug monitoring",
               "population pk","drug dosing","monte carlo simulation"]),
    (5, "5c", ["ceftazidime","cefiderocol","imipenem-cilastatin","meropenem-vaborbactam",
               "ceftolozane","cefepime","aztreonam","new beta-lactam","repurposed antibiotic",
               "clinical trial antibiotic"]),
    (5, "5d", ["antimicrobial stewardship","antibiotic stewardship","prescribing",
               "antibiotic use","rational antibiotic","de-escalation","iv-to-oral"]),
    (5, "5e", ["adverse effect","drug safety","hypersensitivity","allergy antibiotic",
               "nephrotoxicity","hepatotoxicity"]),
    (5, "5f", ["pharmacoepidemiology","cost-effectiveness","pharmacoeconomics"]),
    # ── Category 6: Mycology ──────────────────────────────────────────────────
    (6, "6a", ["aspergillus","candida","cryptococcus","mucor","mould","fungal pathogen",
               "mycobiome","fungal virulence"]),
    (6, "6b", ["invasive fungal infection","invasive aspergillosis","candidaemia",
               "fungal epidemiology","ifd "]),
    (6, "6c", ["diagnostic mycology","fungal diagnosis","galactomannan","beta-glucan",
               "fungal pcr"]),
    (6, "6d", ["antifungal resistance","antifungal susceptibility","azole resistance",
               "echinocandin resistance"]),
    (6, "6e", ["antifungal treatment","antifungal therapy","voriconazole","isavuconazole",
               "amphotericin","caspofungin","anidulafungin","micafungin"]),
    # ── Category 7: Parasitology & Travel ─────────────────────────────────────
    (7, "7a", ["parasite","plasmodium","leishmania","trypanosoma","helminth","schistosoma"]),
    (7, "7b", ["malaria epidemiology","parasitic disease epidemiology","parasite burden"]),
    (7, "7c", ["malaria diagnosis","parasitic diagnosis","microscopy parasite","rapid diagnostic test"]),
    (7, "7d", ["antimalarial","antiparasitic","malaria treatment","parasitic treatment"]),
    (7, "7f", ["travel medicine","traveller","returning traveller","tropical disease",
               "migrant health","refugee health"]),
    # ── Category 8: HAI & IPC ─────────────────────────────────────────────────
    (8, "8a", ["catheter","cvc","central line","clabsi","intravascular catheter"]),
    (8, "8b", ["prosthetic joint infection","pji","implant infection","foreign body infection"]),
    (8, "8c", ["surgical site infection","ssi ","perioperative infection"]),
    (8, "8d", ["ventilator-associated","vap ","hospital-acquired pneumonia","hap "]),
    (8, "8e", ["nosocomial","hospital epidemiology","hospital transmission",
               "infection surveillance","hospital outbreak","decolonisation"]),
    (8, "8f", ["healthcare-associated infection","hai ","clostridioides difficile","cdiff",
               "c. diff","cdiff infection"]),
    (8, "8g", ["infection control intervention","ipc intervention","hand hygiene",
               "contact precautions","isolation"]),
    (8, "8h", ["disinfection","sterilisation","decontamination","medical device"]),
    (8, "8i", ["healthcare worker","hcw infection","occupational infection","staff vaccination"]),
    # ── Category 9: Basic science ─────────────────────────────────────────────
    (9, "9a", ["pathogenesis","virulence factor","toxin","pathogenicity"]),
    (9, "9b", ["host-pathogen","biofilm","animal model","in vitro model"]),
    (9, "9d", ["cellular microbiology","experimental microbiology","in vitro study"]),
    (9, "9e", ["omics","transcriptomics","proteomics","metabolomics","genomics"]),
    (9, "9f", ["immune response","innate immunity","adaptive immunity","t-cell","antibody response"]),
    # ── Category 10: Immunocompromised ────────────────────────────────────────
    (10, "10e", ["neutropaenia","febrile neutropaenia","haematological malignancy",
                 "cancer patient infection","oncology infection"]),
    (10, "10b", ["solid organ transplant","liver transplant","kidney transplant",
                 "lung transplant","organ transplant"]),
    (10, "10c", ["stem cell transplant","hsct","allogeneic transplant","autologous transplant"]),
    (10, "10d", ["car-t","cell therapy","adoptive therapy"]),
    (10, "10f", ["biological therapy","immunosuppression","corticosteroid",
                 "rheumatological","tnf inhibitor","biologic agent"]),
    # ── Category 11: Public health & vaccines ─────────────────────────────────
    (11, "11a", ["vaccine development","vaccine safety","vaccine efficacy","vaccinology",
                 "vaccine immunogenicity"]),
    (11, "11c", ["pneumococcal vaccine","meningococcal vaccine","antibacterial vaccine"]),
    (11, "11e", ["food safety","water safety","environmental health","vector epidemiology"]),
    (11, "11f", ["one health","veterinary","zoonosis","animal reservoir","animal antimicrobial"]),
    (11, "11g", ["pandemic preparedness","health security","global health","biosafety",
                 "biosecurity","climate change infection"]),
    (11, "11h", ["low-resource","lmic","sub-saharan","africa","developing countr",
                 "resource-limited","health equity","vulnerable population"]),
    # ── Category 12: Professional ─────────────────────────────────────────────
    (12, "12a", ["career","professional development","fellowship","training programme"]),
    (12, "12b", ["ethics","publishing","authorship","peer review","academic"]),
    (12, "12c", ["medical education","teaching","curriculum","competency","osce"]),
    (12, "12d", ["diversity","equality","equity","gender","inclusion"]),
    (12, "12e", ["patient","advocacy","patient perspective","patient involvement"]),
]


def keyword_tag(session: dict) -> dict:
    """Keyword-based fallback classification. Returns classification dict."""
    title = session["title"].lower()
    for cat, subcat, keywords in KEYWORD_RULES:
        if any(k in title for k in keywords):
            return {"primary_cat": cat, "primary_subcat": subcat,
                    "secondary_cat": None, "secondary_subcat": None, "confidence": "low"}
    return {"primary_cat": 4, "primary_subcat": "", "secondary_cat": None,
            "secondary_subcat": None, "confidence": "low"}


# ══════════════════════════════════════════════════════════════════════════════
# ANTHROPIC API TAGGER
# ══════════════════════════════════════════════════════════════════════════════

SYSTEM_PROMPT = (
    "You are an expert in infectious diseases and clinical microbiology, "
    "classifying ESCMID conference sessions. Respond ONLY with valid JSON "
    "with no extra text, preamble, or markdown fences."
)


def build_tagging_prompt(session: dict) -> str:
    return f"""Classify this ESCMID Global conference session into the official taxonomy.

Session code : {session['code']}
Session type : {session['type']}
Session title: {session['title']}

Official ESCMID subcategories:
{ALL_SUBCATS_STR}

Return JSON only:
{{
  "primary_cat"    : <integer 1-12>,
  "primary_subcat" : "<e.g. 4j>",
  "secondary_cat"  : <integer 1-12 or null>,
  "secondary_subcat": "<e.g. 5d or null>",
  "confidence"     : "high" | "medium" | "low"
}}

Guidelines:
- primary_cat/subcat = single best fit
- secondary_cat/subcat = second category if clearly relevant, else null
- For sessions spanning multiple domains (e.g. AI applied to AMS) use both fields"""


def load_cache() -> dict:
    return json.loads(CACHE_FILE.read_text()) if CACHE_FILE.exists() else {}


def save_cache(cache: dict):
    CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
    CACHE_FILE.write_text(json.dumps(cache, indent=2))


EXTRACT_CLASSIFY_PROMPT_SYS = (
    "You are an expert in infectious diseases and clinical microbiology. "
    "You are extracting and classifying sessions from a scientific conference programme PDF. "
    "The raw text may have artefacts from a two-column PDF layout — ignore stray text that "
    "belongs to an adjacent column. Respond ONLY with valid JSON, no extra text."
)

def build_extract_classify_prompt(raw_ctx: str) -> str:
    """
    Combined prompt: clean up raw PDF context AND classify in one API call.
    Used with --api-extract flag.
    """
    return f"""Below is raw text extracted from a two-column conference programme PDF.
It may contain artefacts: truncated titles, adjacent-column text bleed-through,
wrapped lines, author names, chair names, or partial content.

<raw_context>
{raw_ctx}
</raw_context>

1. Extract the SESSION TITLE — the descriptive name of the session (not the abstract
   titles of individual talks within it). Clean up any column artefacts.

2. Extract the SESSION TYPE from: Oral Session / Symposium / Educational /
   Meet-the-Expert / Keynote / Late-Breaking / ePoster Flash / Pipeline / Case Session / Other

3. Classify into the best-fit ESCMID official subcategory:
{ALL_SUBCATS_STR}

Return JSON only:
{{
  "title"          : "<clean session title>",
  "type"           : "<session type>",
  "primary_cat"    : <integer 1-12>,
  "primary_subcat" : "<e.g. 4j>",
  "secondary_cat"  : <integer 1-12 or null>,
  "secondary_subcat": "<e.g. 5d or null>",
  "confidence"     : "high" | "medium" | "low"
}}"""


def api_extract_and_classify(session: dict, client, cache: dict) -> dict:
    """
    Uses the API to BOTH clean the session title/type AND classify it.
    More accurate than separate extraction + classification for noisy PDFs.
    Enabled with --api-extract flag.
    """
    raw_ctx = session.get("_raw_ctx", session.get("title", ""))
    cache_key = f"extract|{session['year']}|{session['code']}|{raw_ctx[:80]}"

    if cache_key in cache:
        return cache[cache_key]

    prompt = build_extract_classify_prompt(raw_ctx)

    for attempt in range(MAX_RETRIES):
        try:
            resp = client.messages.create(
                model=ANTHROPIC_MODEL,
                max_tokens=400,
                system=EXTRACT_CLASSIFY_PROMPT_SYS,
                messages=[{"role": "user", "content": prompt}]
            )
            raw = resp.content[0].text.strip()
            raw = re.sub(r'^```json?\s*|\s*```$', '', raw)
            result = json.loads(raw)

            result["title"]           = str(result.get("title") or session["title"])
            result["type"]            = str(result.get("type", session.get("type","Other")))
            result["primary_cat"]     = int(result.get("primary_cat", 4))
            result["primary_subcat"]  = str(result.get("primary_subcat") or "")
            result["secondary_cat"]   = result.get("secondary_cat")
            result["secondary_subcat"]= result.get("secondary_subcat")
            result["confidence"]      = result.get("confidence", "medium")

            cache[cache_key] = result
            return result
        except (json.JSONDecodeError, KeyError, ValueError):
            if attempt == MAX_RETRIES - 1:
                break
            time.sleep(1.0)
        except Exception as exc:
            if attempt == MAX_RETRIES - 1:
                print(f"\n  API error: {exc}")
                break
            time.sleep(API_DELAY * (attempt + 2))

    return {"title": session["title"], "type": session.get("type","Other"),
            "primary_cat": 4, "primary_subcat": "", "secondary_cat": None,
            "secondary_subcat": None, "confidence": "low"}


def api_tag_session(session: dict, client, cache: dict) -> dict:
    """Call Anthropic API to classify a session; use cache if available."""
    key = f"{session['year']}|{session['code']}|{session['title'][:60]}"
    if key in cache:
        return cache[key]

    for attempt in range(MAX_RETRIES):
        try:
            resp = client.messages.create(
                model=ANTHROPIC_MODEL,
                max_tokens=256,
                system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": build_tagging_prompt(session)}]
            )
            raw = resp.content[0].text.strip()
            raw = re.sub(r'^```json?\s*|\s*```$', '', raw)
            result = json.loads(raw)
            result["primary_cat"]     = int(result.get("primary_cat", 4))
            result["primary_subcat"]  = str(result.get("primary_subcat") or "")
            result["secondary_cat"]   = result.get("secondary_cat")
            result["secondary_subcat"]= result.get("secondary_subcat")
            result["confidence"]      = result.get("confidence", "medium")
            cache[key] = result
            return result
        except (json.JSONDecodeError, KeyError, ValueError):
            if attempt == MAX_RETRIES - 1:
                break
            time.sleep(1.0)
        except Exception as exc:
            if attempt == MAX_RETRIES - 1:
                print(f"\n  API error for {session['code']}: {exc}")
                break
            time.sleep(API_DELAY * (attempt + 2))

    fallback = {"primary_cat": 4, "primary_subcat": "", "secondary_cat": None,
                "secondary_subcat": None, "confidence": "low"}
    cache[key] = fallback
    return fallback


def enrich_session(session: dict, tags: dict) -> dict:
    """Merge tagging results into session dict."""
    cat_num  = tags["primary_cat"]
    subcat   = tags.get("primary_subcat", "")
    cat      = CATEGORIES.get(cat_num, CATEGORIES[4])

    session["escmid_cat"]       = cat_num
    session["escmid_cat_name"]  = cat["name"]
    session["escmid_cat_short"] = cat["short"]
    session["escmid_color"]     = cat["color"]
    session["escmid_subcat"]    = subcat
    session["escmid_subcat_name"] = cat["subcats"].get(subcat, "")
    session["escmid_cat2"]      = tags.get("secondary_cat")
    session["escmid_subcat2"]   = tags.get("secondary_subcat") or ""
    session["confidence"]       = tags.get("confidence", "medium")
    if "tags" not in session:
        session["tags"] = []

    # Prefix is AUTHORITATIVE for known prefixes — always overrides context
    # text extraction, which can be wrong due to two-column layout contamination.
    prefix = session.get("code", "")[:2]
    if prefix in PREFIX_TO_TYPE:
        session["type"] = PREFIX_TO_TYPE[prefix]

    return session


def tag_all(sessions: list, use_api: bool, api_extract: bool = False) -> list:
    """
    Tag all sessions.

    use_api=True      — call Anthropic API for classification only
                        (uses pre-extracted title/type from PDF parsing)
    api_extract=True  — call Anthropic API for BOTH extraction clean-up
                        AND classification in one call. Recommended for
                        best accuracy; costs the same as use_api=True.
    """
    cache = load_cache()
    client = None

    if use_api or api_extract:
        if not HAS_ANTHROPIC:
            print("  anthropic not installed — using keyword fallback")
            use_api = api_extract = False
        else:
            api_key = os.environ.get("ANTHROPIC_API_KEY", "")
            if not api_key:
                print("  ANTHROPIC_API_KEY not set — using keyword fallback")
                use_api = api_extract = False
            else:
                client = _anthropic.Anthropic(api_key=api_key)
                if api_extract:
                    print("  Mode: API extraction + classification (--api-extract)")
                else:
                    print("  Mode: API classification only (keyword extraction)")

    iterator = _tqdm(sessions, desc="  Tagging") if HAS_TQDM else sessions
    tagged = []

    for i, session in enumerate(iterator):
        if api_extract and client:
            result = api_extract_and_classify(session, client, cache)
            # Update session title/type with cleaned versions
            session["title"] = result.pop("title", session["title"])
            session["type"]  = result.pop("type",  session["type"])
            tags = result
        elif use_api and client:
            tags = api_tag_session(session, client, cache)
        else:
            tags = keyword_tag(session)

        if (use_api or api_extract) and i % SAVE_CACHE_EVERY == 0:
            save_cache(cache)
            time.sleep(API_DELAY)

        tagged.append(enrich_session(session, tags))

    save_cache(cache)
    return tagged



# ══════════════════════════════════════════════════════════════════════════════
# 94-TAG CLINICAL/METHODOLOGICAL TAXONOMY
# Applied to sessions on top of ESCMID category classification.
# Run with:  python escmid_analyser.py --add-tags
# ══════════════════════════════════════════════════════════════════════════════

TAG_TAXONOMY = {
    "Methods": [
        "methods_Prediction modeling","methods_Large language models",
        "methods_Image recognition","methods_Causal inference",
        "methods_Genomics and sequencing","methods_Systematic review and meta-analysis",
        "methods_Pharmacokinetics and pharmacodynamics","methods_Epidemiological modelling",
    ],
    "Study Design": [
        "study_design_RCT","study_design_Cohort","study_design_Case series",
        "study_design_Modelling","study_design_Basic science","study_design_Review",
    ],
    "ClinMicro": [
        "clinmicro_Bacteriology","clinmicro_Mycobacteriology","clinmicro_Mycology",
        "clinmicro_Virology","clinmicro_Parasitology","clinmicro_Prions",
        "clinmicro_Diagnostics","clinmicro_Antimicrobial resistance",
        "clinmicro_Diagnostic stewardship","clinmicro_Basic science",
    ],
    "Infectious Diseases": [
        "infectiousdz_Antimicrobial stewardship","infectiousdz_Infection prevention and control",
        "infectiousdz_Treatment","infectiousdz_Comparative trials",
        "infectiousdz_Emerging infections","infectiousdz_Epidemiology",
        "infectiousdz_Clinical manifestations","infectiousdz_Case series",
    ],
    "Treatments": [
        "treatments_Antibiotics","treatments_Antivirals","treatments_Antifungals",
        "treatments_Antiparasitics","treatments_Vaccines","treatments_Phage therapy",
        "treatments_Immunotherapy","treatments_Experimental therapy",
    ],
    "Syndromes": [
        "syndrome_Fever of unknown origin","syndrome_Febrile neutropaenia",
        "syndrome_Upper respiratory tract infection","syndrome_Lower respiratory tract infection",
        "syndrome_Urinary tract infection","syndrome_Sepsis and bloodstream infection",
        "syndrome_Endocarditis and cardiovascular infection","syndrome_Intraabdominal infection",
        "syndrome_Hepatitis and liver infection","syndrome_Gastrointestinal infection",
        "syndrome_Meningitis and encephalitis","syndrome_Brain abscess",
        "syndrome_Skin and soft tissue infection","syndrome_Bone and joint infection",
        "syndrome_Ocular infection","syndrome_Sexually transmitted infection",
        "syndrome_HIV and AIDS","syndrome_Surgical site infection",
    ],
    "Special Hosts": [
        "hosts_Immunocompromised","hosts_Transplant","hosts_Oncology and haematology",
        "hosts_Paediatric","hosts_Neonatal","hosts_Elderly",
        "hosts_Pregnancy and maternal","hosts_ICU and critically ill",
    ],
    "AMR Pathogens": [
        "amr_pathogen_ESKAPE","amr_pathogen_MRSA","amr_pathogen_C. difficile",
        "amr_pathogen_Mycobacteria","amr_pathogen_Gonorrhoea","amr_pathogen_Candida",
    ],
    "Microbiome": [
        "microbiome_Gut","microbiome_Respiratory","microbiome_Skin","microbiome_Intervention",
    ],
    "Public Health": [
        "public_health_Surveillance","public_health_Pandemic preparedness",
        "public_health_Policy","public_health_One Health and zoonoses",
        "public_health_Travel medicine","public_health_Outbreak response",
        "public_health_Bioterrorism and biosecurity",
    ],
    "Professional": [
        "professional_Regulatory","professional_Ethics","professional_Bias and equity",
        "professional_Education and training","professional_Guidelines",
        "professional_Career development",
    ],
    "Region": [
        "region_LMICs","region_Europe","region_Africa","region_Asia-Pacific","region_Americas",
    ],
}

ALL_TAGS     = [t for group in TAG_TAXONOMY.values() for t in group]
TAG_SET      = set(ALL_TAGS)
ALL_TAGS_STR = "\n".join(f"  {t}" for t in ALL_TAGS)

TAG_PALETTE = {
    "methods_": "#22d3ee",  "study_design_": "#818cf8",
    "clinmicro_": "#fbbf24", "infectiousdz_": "#34d399",
    "treatments_": "#fb923c","syndrome_": "#f472b6",
    "hosts_": "#a78bfa",    "amr_pathogen_": "#f87171",
    "microbiome_": "#6ee7b7","public_health_": "#86efac",
    "professional_": "#94a3b8","region_": "#4ade80",
}

def tag_color(tag: str) -> str:
    for prefix, color in TAG_PALETTE.items():
        if tag.startswith(prefix):
            return color
    return "#617d9b"


# ── Tagging prompt ────────────────────────────────────────────────────────────
TAG_SYSTEM_PROMPT = (
    "You are an expert in infectious diseases and clinical microbiology. "
    "Tag ESCMID conference sessions using a fixed clinical taxonomy. "
    "Respond ONLY with valid JSON, no preamble or markdown."
)

def build_tag_prompt(session: dict) -> str:
    cat  = CATEGORIES.get(session.get("escmid_cat", 4), CATEGORIES[4])
    sub  = session.get("escmid_subcat", "")
    subn = cat["subcats"].get(sub, "")
    return f"""Assign 1-4 tags to this ESCMID session from the taxonomy below.

Title   : {session["title"]}
Type    : {session["type"]}
Category: {cat["name"]} / {subn}

Taxonomy (all valid tags):
{ALL_TAGS_STR}

Tag priority order: syndrome → hosts → amr_pathogen → clinmicro → treatments → infectiousdz → methods → study_design → public_health → professional → region

Return JSON only:
{{"tags": ["tag1", "tag2"]}}

Rules:
- Maximum 4 tags, minimum 1
- Prefer specific over generic (e.g. syndrome_Sepsis over infectiousdz_Treatment)
- Only use tags from the list above exactly as written"""


def api_add_tags(session: dict, client, cache: dict) -> list:
    """Call API to assign 94-taxonomy tags to one session."""
    key = f"tags94|{session['year']}|{session['code']}|{session['title'][:60]}"
    if key in cache:
        return cache[key]
    # Fallback: scan for any cached key matching this year+code
    # (titles may have changed slightly after extraction fixes)
    prefix = f"tags94|{session['year']}|{session['code']}|"
    for k, v in cache.items():
        if k.startswith(prefix):
            cache[key] = v   # store under new key for future runs
            return v

    for attempt in range(MAX_RETRIES):
        try:
            resp = client.messages.create(
                model=ANTHROPIC_MODEL, max_tokens=150,
                system=TAG_SYSTEM_PROMPT,
                messages=[{"role": "user", "content": build_tag_prompt(session)}]
            )
            if not resp.content:
                raise ValueError("empty response")
            raw = resp.content[0].text.strip()
            raw = re.sub(r"^```json?\s*|\s*```$", "", raw)
            result = json.loads(raw)
            tags = [t for t in result.get("tags", []) if t in TAG_SET][:4]
            cache[key] = tags
            return tags
        except (json.JSONDecodeError, KeyError, ValueError):
            if attempt == MAX_RETRIES - 1: break
            time.sleep(1.0)
        except Exception as exc:
            if attempt == MAX_RETRIES - 1:
                print(f"\n  Tag API error for {session['code']}: {exc}")
                break
            time.sleep(API_DELAY * (attempt + 2))

    cache[key] = []
    return []


def run_tag_pass(sessions: list) -> list:
    """
    Add 94-taxonomy tags to all sessions.
    Designed to be run AFTER categorisation (uses cached ESCMID categories).
    Tags are cached separately so this pass can be re-run independently.
    """
    cache = load_cache()
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not (HAS_ANTHROPIC and api_key):
        print("  ANTHROPIC_API_KEY not set — skipping tag pass")
        return sessions

    client = _anthropic.Anthropic(api_key=api_key)
    already = sum(1 for s in sessions if s.get("tags"))
    print(f"  {already}/{len(sessions)} sessions already have tags (cached)")

    iterator = _tqdm(sessions, desc="  Tagging (94 tags)") if HAS_TQDM else sessions
    for i, session in enumerate(iterator):
        if not session.get("tags"):
            session["tags"] = api_add_tags(session, client, cache)
            if i % SAVE_CACHE_EVERY == 0:
                save_cache(cache)
            time.sleep(API_DELAY)

    save_cache(cache)
    return sessions

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL GENERATION
# ══════════════════════════════════════════════════════════════════════════════

def _border():
    thin = Side(style="thin", color="DDDDDD")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _hdr_cell(ws, row, col, value, fill_hex="1F4E79"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill    = PatternFill("solid", start_color=fill_hex)
    c.font    = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border  = _border()
    return c

def _data_cell(ws, row, col, value, fill_hex="FFFFFF"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill    = PatternFill("solid", start_color=fill_hex)
    c.font    = Font(name="Arial", size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.border  = _border()
    return c


def write_overview_sheet(ws, sessions: list, years: list):
    """Summary matrix: ESCMID category × year."""
    _hdr_cell(ws, 1, 1, "ESCMID Category")
    for j, yr in enumerate(years, 2):
        _hdr_cell(ws, 1, j, yr)
    _hdr_cell(ws, 1, len(years)+2, "TOTAL")
    ws.row_dimensions[1].height = 30

    counts = defaultdict(lambda: defaultdict(int))
    for s in sessions:
        counts[s["escmid_cat"]][s["year"]] += 1

    # Build subcategory counts
    subcat_counts = defaultdict(lambda: defaultdict(int))
    for s in sessions:
        if s.get("escmid_subcat"):
            subcat_counts[(s["escmid_cat"], s["escmid_subcat"])][s["year"]] += 1

    current_row = 2
    for cat_num, cat in CATEGORIES.items():
        h = cat["color"].lstrip("#") + "44"
        # Category header row
        _data_cell(ws, current_row, 1, f"{cat_num}. {cat['name']}", h)
        ws.cell(current_row, 1).font = Font(bold=True, name="Arial", size=10)
        total = 0
        for j, yr in enumerate(years, 2):
            v = counts[cat_num].get(yr, 0)
            total += v
            _data_cell(ws, current_row, j, v or "", h)
            ws.cell(current_row, j).alignment = Alignment(horizontal="center")
        _data_cell(ws, current_row, len(years)+2, total, h)
        ws.cell(current_row, len(years)+2).font = Font(bold=True, name="Arial", size=10)
        current_row += 1
        # Subcategory detail rows
        light_h = cat["color"].lstrip("#") + "18"
        for sub_code, sub_name in cat["subcats"].items():
            sub_total = sum(subcat_counts[(cat_num, sub_code)].values())
            if sub_total == 0:
                continue
            _data_cell(ws, current_row, 1, f"    {sub_code}: {sub_name}", light_h)
            ws.cell(current_row, 1).font = Font(italic=True, name="Arial", size=9)
            st = 0
            for j, yr in enumerate(years, 2):
                v = subcat_counts[(cat_num, sub_code)].get(yr, 0)
                st += v
                _data_cell(ws, current_row, j, v or "", light_h)
                ws.cell(current_row, j).alignment = Alignment(horizontal="center")
            _data_cell(ws, current_row, len(years)+2, st or "", light_h)
            current_row += 1

    ws.column_dimensions["A"].width = 46
    for j in range(2, len(years)+3):
        ws.column_dimensions[ws.cell(1, j).column_letter].width = 10
    ws.freeze_panes = "B2"


def write_category_sheet(ws, sessions: list, cat_num: int, years: list):
    """One sheet per ESCMID category."""
    cat = CATEGORIES[cat_num]
    col_hex = cat["color"].lstrip("#")

    headers = ["Year","Code","Type","Session Title",
               "ESCMID Subcategory","Subcategory Name",
               "Secondary Cat","Tags","Confidence"]
    for j, h in enumerate(headers, 1):
        _hdr_cell(ws, 1, j, h)
    ws.row_dimensions[1].height = 30

    # Sort by year descending
    rows = sorted(sessions, key=lambda x: x["year"], reverse=True)

    for i, s in enumerate(rows, 2):
        fill = col_hex + ("33" if i % 2 == 0 else "11")
        cat2 = CATEGORIES.get(s.get("escmid_cat2"))
        cat2_str = (f"{s['escmid_cat2']} – {cat2['short']}"
                    if cat2 else "")
        tag_str = "; ".join(s.get("tags", []))
        vals = [s["year"], s["code"], s["type"], s["title"],
                s.get("escmid_subcat",""), s.get("escmid_subcat_name",""),
                cat2_str, tag_str, s.get("confidence","")]
        for j, v in enumerate(vals, 1):
            _data_cell(ws, i, j, v, fill)

    widths = [8, 10, 18, 60, 10, 40, 22, 50, 10]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1,j).column_letter].width = w
    ws.freeze_panes = "A2"


def generate_excel(sessions: list, path: Path):
    if not HAS_OPENPYXL:
        print("  openpyxl not installed — skipping Excel output")
        return

    years = sorted(set(s["year"] for s in sessions))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Overview"
    write_overview_sheet(ws, sessions, years)

    for cat_num, cat in CATEGORIES.items():
        cat_sessions = [s for s in sessions if s.get("escmid_cat") == cat_num]
        sheet_name = f"{cat_num}. {cat['short']}"[:31]
        write_category_sheet(wb.create_sheet(sheet_name), cat_sessions, cat_num, years)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"  Excel  → {path}")


# ══════════════════════════════════════════════════════════════════════════════
# HTML DASHBOARD GENERATION
# ══════════════════════════════════════════════════════════════════════════════

def _extract_chairs(raw_ctx: str) -> str:
    """Extract chair names from raw PDF context."""
    lines = raw_ctx.split("\n")
    chairs = []
    capture = False
    for line in lines:
        line = line.strip()
        if re.match(r"^Chairs?\b", line, re.I):
            capture = True
            rest = re.sub(r"^Chairs?\s*", "", line, flags=re.I).strip()
            if rest and not re.match(r"^\t", rest):
                chairs.append(rest)
            continue
        if capture:
            if not line or re.match(r"^(Co-organised|[A-Z][a-z]+ \d{2})", line):
                break
            # Strip leading tab/whitespace artefacts
            clean = re.sub(r"^[\t ]+", "", line)
            if clean and len(clean) > 3:
                chairs.append(clean)
            if len(chairs) >= 3:
                break
    # Clean partial names (cut off by column width)
    return "; ".join(c.rstrip(",") for c in chairs if len(c) > 4)


def _extract_time_hall(raw_ctx: str) -> str:
    """Extract time and hall from raw PDF context."""
    m = re.search(r"(\d{2}:\d{2}\s*[-–]\s*\d{2}:\d{2})", raw_ctx)
    time_str = m.group(1) if m else ""
    h = re.search(r"(Hall\s+[\w\d]+|Arena\s+[\w\d]+|Room\s+[\w\d]+)", raw_ctx, re.I)
    hall_str = h.group(1) if h else ""
    parts = [p for p in [time_str, hall_str] if p]
    return "  ·  ".join(parts)



def _clean_person_name(raw: str) -> str:
    """Normalise a raw chair/speaker string to a clean display name."""
    name = raw.replace('*', '').strip()
    # Remove partial city/country fragments that leaked through column truncation
    # e.g. "Kathrin Blum (Zurich, Swit" → "Kathrin Blum"
    name = re.sub(r'\s*\([^)]*$', '', name).strip()    # unclosed parenthesis (truncated)
    name = re.sub(r'\s*\([^)]*\)', '', name).strip()   # closed (City, Country)
    name = name.rstrip('.,;').strip()
    # Reject obvious non-name artefacts
    if name.endswith(')') or re.match(r'^[A-Z][a-z]+\)', name):
        return ''
    return name


def _is_initial_name(name: str) -> bool:
    """True if name is just initials like 'A. B.' — too ambiguous to index."""
    if re.match(r'^[A-Z][a-z]?\.(?:\s+[A-Z][a-z]?\.)*$', name.strip()):
        return True
    # Also reject single-word fragments (likely country/city artefacts)
    words = name.split()
    if len(words) == 1 and not re.match(r'^[A-Z][a-z]{2,}', words[0]):
        return True
    return False


def build_people_index(sessions: list) -> list:
    """
    Build a cross-year index of chairs and speakers.
    Returns list of person dicts sorted by number of distinct years active.
    Only includes people whose full name could be resolved (≥2 words).
    """
    people: dict = defaultdict(lambda: {"display": "", "apps": []})

    for s in sessions:
        year  = s["year"]
        code  = s["code"]
        title = s["title"][:80]
        cat   = s.get("escmid_cat", 4)
        color = s.get("escmid_color", "#617d9b")

        # Chairs
        for raw in re.split(r';', s.get("chairs", "") or ""):
            name = _clean_person_name(raw)
            if len(name.split()) >= 2 and not _is_initial_name(name):
                key = name.lower()
                if not people[key]["display"]:
                    people[key]["display"] = name
                people[key]["apps"].append({
                    "year": year, "code": code, "title": title,
                    "role": "Chair", "cat": cat, "color": color,
                    "talk_code": "", "talk_title": "",
                })

        # Speakers from individual talks
        for talk in s.get("talks", []):
            for raw in re.split(r';', talk.get("speakers", "") or ""):
                name = _clean_person_name(raw)
                if len(name.split()) >= 2 and not _is_initial_name(name):
                    key = name.lower()
                    if not people[key]["display"]:
                        people[key]["display"] = name
                    people[key]["apps"].append({
                        "year": year, "code": code, "title": title,
                        "role": "Speaker", "cat": cat, "color": color,
                        "talk_code": talk.get("code", ""),
                        "talk_title": talk.get("title", "")[:80],
                    })

    # Flatten, deduplicate appearances, sort
    result = []
    for data in people.values():
        if not data["apps"]:
            continue
        years = sorted(set(a["year"] for a in data["apps"]))
        result.append({
            "name":          data["display"],
            "years":         years,
            "count":         len(data["apps"]),
            "chair_count":   sum(1 for a in data["apps"] if a["role"] == "Chair"),
            "speaker_count": sum(1 for a in data["apps"] if a["role"] == "Speaker"),
            "apps":          sorted(data["apps"], key=lambda x: x["year"], reverse=True)[:30],
        })

    # Sort: most years active first, then total appearances
    result.sort(key=lambda x: (-len(x["years"]), -x["count"]))
    # Cap at 2000 to keep dashboard size manageable
    return result[:2000]

def build_dashboard_data(sessions: list) -> dict:
    years = sorted(set(s["year"] for s in sessions))

    cat_year  = defaultdict(lambda: defaultdict(int))
    cat_type  = defaultdict(lambda: defaultdict(int))
    subcat_year = defaultdict(lambda: defaultdict(int))

    for s in sessions:
        cn = s.get("escmid_cat", 4)
        cat_year[cn][s["year"]] += 1
        cat_type[cn][s["type"]] += 1
        sc = s.get("escmid_subcat", "")
        if sc:
            subcat_year[sc][s["year"]] += 1

    # Build row list for explore + network
    rows = []
    for s in sessions:
        rows.append({
            "id"        : len(rows),
            "year"      : s["year"],
            "code"      : s["code"],
            "type"      : s["type"],
            "title"     : s["title"],
            "cat"       : s.get("escmid_cat", 4),
            "cat_short" : s.get("escmid_cat_short", ""),
            "cat_name"  : s.get("escmid_cat_name", ""),
            "subcat"    : s.get("escmid_subcat", ""),
            "subcat_name": s.get("escmid_subcat_name", ""),
            "cat2"      : s.get("escmid_cat2"),
            "color"     : s.get("escmid_color", "#617d9b"),
            "confidence": s.get("confidence", "medium"),
            "tags"      : s.get("tags", []),
            "tag_colors": [tag_color(t) for t in s.get("tags", [])],
            "chairs"    : _extract_chairs(s.get("_raw_ctx", "")),
            "time_hall" : _extract_time_hall(s.get("_raw_ctx", "")),
            "talks"     : s.get("talks", []),
        })

    # Network nodes: exclude high-volume/low-signal session types for performance
    _NET_EXCLUDE = {"ePoster Flash", "IS / Integrated", "Other"}
    net_rows = [r for r in rows if r["type"] not in _NET_EXCLUDE]
    for i, r in enumerate(net_rows):
        r["net_id"] = i

    edges = []
    seen  = set()
    for i, r1 in enumerate(net_rows):
        for j, r2 in enumerate(net_rows):
            if i >= j: continue
            # Shared tags (primary signal); fall back to shared subcategory
            t1, t2 = set(r1.get("tags",[])), set(r2.get("tags",[]))
            shared_tags = len(t1 & t2)
            if shared_tags > 0:
                weight = shared_tags
            elif (r1["cat"] == r2["cat"] and r1.get("subcat")
                  and r1.get("subcat") == r2.get("subcat")):
                weight = 1
            else:
                weight = 0
            if weight > 0:
                key = (min(i,j), max(i,j))
                if key not in seen and len(edges) < 4000:
                    seen.add(key)
                    edges.append({"s": i, "t": j, "w": weight})

    # Tag × year counts for heatmap
    tag_year = defaultdict(lambda: defaultdict(int))
    for s in sessions:
        for t in s.get("tags", []):
            tag_year[t][s["year"]] += 1

    people = build_people_index(sessions)

    return {
        "rows"       : rows,
        "people"     : people,
        "net_rows"   : net_rows,
        "net_edges"  : edges,
        "years"      : years,
        "year_total" : dict(Counter(s["year"] for s in sessions)),
        "tag_year"   : {t: dict(v) for t,v in tag_year.items()},
        "tag_taxonomy": TAG_TAXONOMY,
        "tag_palette" : TAG_PALETTE,
        "cat_year"   : {str(k): dict(v) for k, v in cat_year.items()},
        "cat_type"   : {str(k): dict(v) for k, v in cat_type.items()},
        "subcat_year": {k: dict(v) for k, v in subcat_year.items()},
        "categories" : {
            str(k): {"name": v["name"], "short": v["short"],
                     "color": v["color"], "subcats": v["subcats"]}
            for k, v in CATEGORIES.items()
        },
    }



def generate_dashboard_html(data_json: str) -> str:
    """Return complete HTML for the dashboard with all enhancements."""

    YEAR_COLORS = {
        "2021":"#94a3b8","2022":"#7dd3fc","2023":"#6ee7b7",
        "2024":"#fbbf24","2025":"#4fc3f7","2026":"#818cf8",
    }
    yc_js = json.dumps(YEAR_COLORS)

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>ESCMID EPC Copilot</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Sora:wght@300;400;600;700&family=JetBrains+Mono:wght@400;600&display=swap" rel="stylesheet">
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/d3/7.8.5/d3.min.js"></script>
<style>
:root{{--bg:#080e1a;--surf:#0f1d2e;--surf2:#162336;--bdr:#1e3450;--text:#c8d8ec;--dim:#617d9b;--acc:#4fc3f7}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{background:var(--bg);color:var(--text);font-family:'Sora',sans-serif;font-size:13px;line-height:1.6;height:100vh;display:flex;flex-direction:column;overflow:hidden}}
header{{padding:12px 32px;border-bottom:1px solid var(--bdr);background:linear-gradient(135deg,#080e1a,#0a1830,#0c1e35);flex-shrink:0;display:flex;align-items:center;gap:20px;flex-wrap:wrap}}
.logo{{font-size:17px;font-weight:700;color:#e8f4ff;letter-spacing:-.02em}}.logo span{{color:var(--acc)}}
.hstats{{display:flex;gap:20px;margin-left:auto;flex-wrap:wrap}}
.stat{{display:flex;flex-direction:column;gap:1px}}
.sn{{font-family:'JetBrains Mono',monospace;font-size:20px;font-weight:600;line-height:1}}
.sl{{font-size:10px;color:var(--dim);text-transform:uppercase;letter-spacing:.1em}}
.tab-bar{{display:flex;background:rgba(8,14,26,.95);border-bottom:1px solid var(--bdr);flex-shrink:0;padding:0 32px;overflow-x:auto}}
.tab-bar::-webkit-scrollbar{{height:0}}
.tab{{display:flex;align-items:center;gap:6px;padding:10px 14px;font-size:11px;font-weight:600;color:var(--dim);background:none;border:none;border-bottom:2px solid transparent;cursor:pointer;letter-spacing:.04em;text-transform:uppercase;white-space:nowrap;transition:color .18s,border-color .18s;font-family:'Sora',sans-serif}}
.tab:hover{{color:var(--text)}}.tab.active{{color:var(--acc);border-color:var(--acc)}}
.content{{flex:1;overflow:hidden}}
.pane{{display:none;height:100%;overflow-y:auto;padding:24px 32px}}
.pane.active{{display:block}}
.pane::-webkit-scrollbar{{width:5px}}.pane::-webkit-scrollbar-thumb{{background:var(--bdr);border-radius:3px}}
.ptitle{{font-size:15px;font-weight:700;color:#e8f4ff;margin-bottom:6px}}
.pdesc{{font-size:12px;color:var(--dim);margin-bottom:18px;max-width:700px;line-height:1.7}}
.g2{{display:grid;grid-template-columns:1fr 1fr;gap:16px}}
.g2h{{display:grid;grid-template-columns:1fr 1fr;gap:16px;align-items:start}}
.card{{background:var(--surf);border:1px solid var(--bdr);border-radius:8px;padding:16px 20px}}
.ctitle{{font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:.1em;color:var(--dim);margin-bottom:12px}}
/* Heatmaps */
#heat-wrap-l,#heat-wrap-r{{overflow:auto;max-height:500px;font-family:'JetBrains Mono',monospace;font-size:10px}}
.hm-cell{{height:24px;border-radius:3px;display:flex;align-items:center;justify-content:center;font-size:9px;cursor:default;transition:transform .12s}}
.hm-cell:hover{{transform:scale(1.12);z-index:2}}
.hm-lbl{{font-size:10px;color:var(--dim);display:flex;align-items:center;padding-right:6px;white-space:nowrap;height:24px}}
.hm-yr{{font-family:'JetBrains Mono',monospace;font-size:9px;text-align:center;height:18px;display:flex;align-items:center;justify-content:center;font-weight:600;color:var(--acc)}}
/* Network */
#net-card{{padding:0;overflow:hidden;display:flex;flex-direction:column;height:calc(100vh - 185px);min-height:460px}}
#net-ctrl{{display:flex;flex-wrap:wrap;gap:6px;align-items:center;padding:9px 14px;border-bottom:1px solid var(--bdr);background:var(--surf2);flex-shrink:0}}
.cl{{font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:.1em;color:var(--dim);white-space:nowrap}}
.cg{{display:flex;gap:4px;flex-wrap:wrap}}
.cb{{font-family:'JetBrains Mono',monospace;font-size:9.5px;padding:3px 10px;border-radius:12px;border:1px solid var(--bdr);background:var(--surf);color:var(--dim);cursor:pointer;transition:all .15s;white-space:nowrap}}
.cb:hover{{border-color:rgba(255,255,255,.3);color:#e8f4ff}}.cb.on{{color:#fff}}
.csep{{width:1px;height:18px;background:var(--bdr);margin:0 3px}}
#net-body{{display:flex;flex:1;overflow:hidden;min-height:0}}
#net-svg-wrap{{flex:1;overflow:hidden}}
#net-svg-wrap svg{{width:100%;height:100%;display:block;cursor:grab}}
#net-svg-wrap svg:active{{cursor:grabbing}}
.lnk{{stroke-opacity:.1}}
#dp{{width:250px;flex-shrink:0;border-left:1px solid var(--bdr);background:var(--surf2);overflow-y:auto;padding:12px;display:flex;flex-direction:column;gap:7px}}
.de{{color:var(--dim);font-size:11px;text-align:center;margin:auto;line-height:2.2;opacity:.7}}
.dy{{font-family:'JetBrains Mono',monospace;font-size:11px;font-weight:600;margin-bottom:2px}}
.dt{{font-size:12px;color:#e8f4ff;line-height:1.5;font-weight:600}}
.ds{{font-size:10px;color:var(--dim);margin-top:2px;line-height:1.4}}
.dtags{{display:flex;flex-wrap:wrap;gap:3px;margin-top:6px}}
.dtag{{font-size:9px;font-family:'JetBrains Mono',monospace;padding:2px 7px;border-radius:10px}}
.dnbl{{font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:.1em;color:var(--dim);margin:7px 0 4px}}
.nbc{{font-size:10px;color:var(--dim);background:var(--surf);border:1px solid var(--bdr);border-radius:4px;padding:3px 8px;cursor:pointer;transition:border-color .15s;line-height:1.4;margin-bottom:3px;display:block;text-align:left;width:100%;font-family:'Sora',sans-serif}}
.nbc:hover{{border-color:var(--acc);color:var(--text)}}
/* Explore */
.filters{{display:flex;flex-wrap:wrap;gap:8px;margin-bottom:12px;align-items:center}}
.fb{{font-family:'JetBrains Mono',monospace;font-size:9.5px;padding:4px 10px;border-radius:18px;border:1px solid var(--bdr);background:var(--surf2);color:var(--dim);cursor:pointer;transition:all .15s;white-space:nowrap}}
.fb:hover{{border-color:var(--acc);color:var(--acc)}}.fb.on{{background:rgba(79,195,247,.15);border-color:var(--acc);color:var(--acc)}}
.sb{{flex:1;min-width:160px;max-width:300px;background:var(--surf2);border:1px solid var(--bdr);color:var(--text);padding:5px 12px;border-radius:18px;font-size:12px;font-family:'Sora',sans-serif;outline:none;transition:border-color .2s}}
.sb:focus{{border-color:var(--acc)}}.sb::placeholder{{color:var(--dim)}}
.tc{{font-family:'JetBrains Mono',monospace;font-size:11px;color:var(--dim);margin-left:auto}}
.tl{{display:flex;flex-direction:column;gap:4px;max-height:calc(100vh - 300px);overflow-y:auto;padding-right:3px}}
.tl::-webkit-scrollbar{{width:4px}}.tl::-webkit-scrollbar-thumb{{background:var(--bdr);border-radius:2px}}
.ti{{background:var(--surf2);border:1px solid var(--bdr);border-radius:6px;padding:7px 12px;display:grid;grid-template-columns:36px 1fr auto;gap:5px 10px;align-items:start;transition:border-color .15s}}
.ti:hover{{border-color:rgba(79,195,247,.3)}}
.ty{{font-family:'JetBrains Mono',monospace;font-size:10px;color:var(--acc);font-weight:600;line-height:1.8}}
.tt2{{font-size:12px;color:var(--text);line-height:1.4}}
.ts{{font-size:10px;color:var(--dim);margin-top:2px}}
.tgs{{display:flex;flex-wrap:wrap;gap:3px;justify-content:flex-end;max-width:220px}}
.tp{{font-size:9px;padding:2px 6px;border-radius:10px;white-space:nowrap;font-family:'JetBrains Mono',monospace}}
/* Gaps */
.gg{{display:grid;grid-template-columns:repeat(auto-fill,minmax(190px,1fr));gap:10px}}
.gc{{background:var(--surf2);border:1px solid var(--bdr);border-radius:6px;padding:12px 14px;border-left:3px solid}}
.gc.low{{border-left-color:#f87171}}.gc.med{{border-left-color:#fbbf24}}
.gct{{font-family:'JetBrains Mono',monospace;font-size:9.5px;color:var(--dim);margin-bottom:3px}}
.gcn{{font-size:20px;font-weight:700;font-family:'JetBrains Mono',monospace}}
.gcn.low{{color:#f87171}}.gcn.med{{color:#fbbf24}}
.gcx{{font-size:11px;color:var(--dim);margin-top:3px}}
/* Trend per-category */
.cat-trend-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(380px,1fr));gap:14px;margin-top:16px}}
.cat-trend-card{{background:var(--surf2);border:1px solid var(--bdr);border-radius:7px;padding:14px 16px}}
.cat-trend-title{{font-size:11px;font-weight:600;margin-bottom:10px;display:flex;align-items:center;gap:8px}}
.cat-dot{{width:10px;height:10px;border-radius:50%;flex-shrink:0}}
/* Explore dropdowns */
.esel{{background:var(--surf2);border:1px solid var(--bdr);color:var(--text);padding:5px 10px;border-radius:18px;font-size:11px;font-family:'Sora',sans-serif;outline:none;cursor:pointer;transition:border-color .2s;min-width:130px;max-width:200px}}
.esel:focus,.esel:hover{{border-color:var(--acc)}}
.esel option{{background:var(--surf2);color:var(--text)}}
/* Tooltip */
.tt{{position:fixed;background:#0a1525;border:1px solid var(--acc);border-radius:6px;padding:7px 11px;font-size:11px;pointer-events:none;z-index:9999;max-width:260px;display:none;color:var(--text);box-shadow:0 8px 24px rgba(0,0,0,.6)}}
.tt.on{{display:block}}
@media(max-width:800px){{
  header{{padding:12px 16px}}.tab-bar{{padding:0 12px}}.pane{{padding:18px 16px}}
  .g2,.g2h{{grid-template-columns:1fr}}
  #net-body{{flex-direction:column}}#dp{{width:100%;height:180px;border-left:none;border-top:1px solid var(--bdr)}}
  #net-card{{height:auto}}
}}
</style>
</head>
<body>
<div class="tt" id="tt"></div>
<header>
  <div class="logo">ESCMID EPC <span>Copilot</span></div>
  <div class="hstats">
    <div class="stat"><span class="sn" style="color:var(--acc)" id="hd-s">—</span><span class="sl">Sessions</span></div>
    <div class="stat"><span class="sn" style="color:#fbbf24" id="hd-y">—</span><span class="sl">Years</span></div>
    <div class="stat"><span class="sn" style="color:#34d399">12</span><span class="sl">Categories</span></div>
    <div class="stat"><span class="sn" style="color:#a78bfa">94</span><span class="sl">Tags</span></div>
  </div>
</header>
<div class="tab-bar">
  <button class="tab active" data-tab="overview">📊 Overview</button>
  <button class="tab" data-tab="trends">📈 Trends</button>
  <button class="tab" data-tab="heatmap">🔥 Heatmap</button>
  <button class="tab" data-tab="network">🕸 Network</button>
  <button class="tab" data-tab="explore">🔍 Explore</button>
  <button class="tab" data-tab="gaps">🎯 Gaps</button>
  <button class="tab" data-tab="people">👥 People</button>
</div>
<div class="content">
  <div class="pane active" id="pane-overview">
    <div class="ptitle">Programme Overview</div>
    <p class="pdesc">Sessions per year by ESCMID category. Click a category in the legend to <strong>isolate</strong> it; click again to restore all.</p>
    <div class="g2">
      <div class="card"><div class="ctitle">Sessions by category per year</div><div style="position:relative;height:270px"><canvas id="chtOv"></canvas></div></div>
      <div class="card"><div class="ctitle">Session type mix per year</div><div style="position:relative;height:270px"><canvas id="chtTy"></canvas></div></div>
    </div>
  </div>
  <div class="pane" id="pane-trends">
    <div class="ptitle">Trends</div>
    <p class="pdesc">Overall category trends over time, plus per-category subcategory breakdown. Click legend items to isolate.</p>
    <div class="card"><div class="ctitle">All categories over time</div><div style="position:relative;height:280px"><canvas id="chtAll"></canvas></div></div>
    <div class="cat-trend-grid" id="cat-trend-grid"></div>
  </div>
  <div class="pane" id="pane-heatmap">
    <div class="ptitle">Heatmaps</div>
    <p class="pdesc">Left: ESCMID subcategories × year. Right: 94-tag taxonomy × year. Hover cells for counts.</p>
    <div class="g2h">
      <div class="card"><div class="ctitle">ESCMID Subcategories</div><div id="heat-wrap-l"><div id="hmc-l"></div></div></div>
      <div class="card"><div class="ctitle">94-Tag Taxonomy</div><div id="heat-wrap-r"><div id="hmc-r"></div></div></div>
    </div>
  </div>
  <div class="pane" id="pane-network">
    <div class="ptitle" style="margin-bottom:5px">Clustering Network</div>
    <p class="pdesc">Nodes = sessions. Edge weight = shared tags (thicker = more in common). Coloured by ESCMID category. <strong>Click</strong> to inspect · <strong>scroll</strong> to zoom · <strong>drag</strong> to pan.</p>
    <div class="card" id="net-card">
      <div id="net-ctrl">
        <span class="cl">Year</span><div class="cg" id="net-yr"></div>
        <div class="csep"></div>
        <span class="cl">Category</span><div class="cg" id="net-cat"></div>
        <div class="csep"></div>
        <button class="cb on" id="btn-edges">Edges on</button>
        <button class="cb on" id="btn-cluster">Cluster on</button>
        <button class="cb" id="btn-reset" style="margin-left:auto">↺ Reset</button>
      </div>
      <div id="net-body">
        <div id="net-svg-wrap"><svg id="net-svg"></svg></div>
        <div id="dp"><div class="de">Click any node<br>to see details</div></div>
      </div>
    </div>
  </div>
  <div class="pane" id="pane-explore">
    <div class="ptitle">Explore All Sessions</div>
    <p class="pdesc">Search and filter sessions. Click any row for details. Sorted by year descending.</p>
    <div class="card">
      <div class="filters">
        <input type="text" class="sb" id="search" placeholder="Search session titles…">
        <select class="esel" id="ef-yr-sel"><option value="">All years</option></select>
        <select class="esel" id="ef-cat-sel"><option value="">All categories</option></select>
        <select class="esel" id="ef-sub-sel"><option value="">All subcategories</option></select>
        <select class="esel" id="ef-tag-sel"><option value="">All tags</option></select>
        <button class="fb" id="btn-show-all" style="margin-left:4px" title="Toggle ePosters, IS and Other sessions">ePosters hidden</button>
        <span class="tc" id="etc">— sessions</span>
      </div>
      <div class="tl" id="tlist"></div>
    </div>
  </div>
  <!-- Session detail modal -->
  <div id="modal-overlay" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.6);z-index:2000;align-items:center;justify-content:center" onclick="if(event.target===this)closeModal()">
    <div id="modal-box" style="background:var(--surf);border:1px solid var(--bdr);border-radius:10px;padding:24px 28px;max-width:640px;width:90%;max-height:80vh;overflow-y:auto;position:relative">
      <button onclick="closeModal()" style="position:absolute;top:12px;right:14px;background:none;border:none;color:var(--dim);font-size:18px;cursor:pointer;line-height:1">✕</button>
      <div id="modal-body"></div>
    </div>
  </div>
  <div class="pane" id="pane-people">
    <div class="ptitle">People</div>
    <p class="pdesc">Chairs and speakers across all years. Sourced from session chair listings and extracted talk speaker lines — quality improves after <code style="font-size:11px;background:var(--surf2);padding:1px 5px;border-radius:3px">--extract-talks</code>. Click any row to see their full session history.</p>
    <div class="card">
      <div class="filters">
        <input type="text" class="sb" id="p-search" placeholder="Search by name…">
        <select class="esel" id="p-yr"><option value="">All years</option></select>
        <select class="esel" id="p-role"><option value="">All roles</option><option value="Chair">Chair only</option><option value="Speaker">Speaker only</option></select>
        <select class="esel" id="p-cat"><option value="">All categories</option></select>
        <span class="tc" id="p-count">— people</span>
      </div>
      <div class="tl" id="p-list" style="max-height:calc(100vh - 310px)"></div>
    </div>
  </div>
  <!-- Person detail modal -->
  <div id="person-modal" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.6);z-index:2000;align-items:center;justify-content:center" onclick="if(event.target===this)closePersonModal()">
    <div id="person-box" style="background:var(--surf);border:1px solid var(--bdr);border-radius:10px;padding:22px 26px;max-width:680px;width:92%;max-height:82vh;overflow-y:auto;position:relative">
      <button onclick="closePersonModal()" style="position:absolute;top:10px;right:13px;background:none;border:none;color:var(--dim);font-size:18px;cursor:pointer">✕</button>
      <div id="person-body"></div>
    </div>
  </div>
  <div class="pane" id="pane-gaps">
    <div class="ptitle">Under-represented Areas</div>
    <p class="pdesc">ESCMID subcategories and tags with the fewest sessions — potential gaps for 2027 proposals.</p>
    <div class="g2" style="margin-bottom:16px">
      <div class="card"><div class="ctitle">ESCMID subcategory gaps</div><div class="gg" id="gg-sub"></div></div>
      <div class="card"><div class="ctitle">Tag gaps</div><div class="gg" id="gg-tag"></div></div>
    </div>
  </div>
</div>

<script>
const D={data_json};
const ROWS=D.rows,YEARS=D.years,CAT_YEAR=D.cat_year,CATS=D.categories;
const YEAR_TOTAL=D.year_total,TAG_YEAR=D.tag_year||{{}};
const TAG_TAX=D.tag_taxonomy||{{}},TAG_PAL=D.tag_palette||{{}};
const YCOLORS={yc_js};

function h2r(hex,a){{const r=parseInt(hex.slice(1,3),16),g=parseInt(hex.slice(3,5),16),b=parseInt(hex.slice(5,7),16);return `rgba(${{r}},${{g}},${{b}},${{a}})`}}
function catColor(cn){{return (CATS[cn]||{{}}).color||'#617d9b'}}
function tagColor(tag){{for(const[p,c] of Object.entries(TAG_PAL)){{if(tag.startsWith(p))return c;}}return '#617d9b';}}

document.getElementById('hd-s').textContent=D.net_rows.length.toLocaleString();
document.getElementById('hd-y').textContent=YEARS.length;

// Tab switching
let netInit=false;
document.querySelectorAll('.tab').forEach(b=>b.addEventListener('click',()=>{{
  document.querySelectorAll('.tab,.pane').forEach(x=>x.classList.remove('active'));
  b.classList.add('active');
  document.getElementById('pane-'+b.dataset.tab).classList.add('active');
  if(b.dataset.tab==='network'){{if(!netInit){{buildNetwork();netInit=true;}}else{{window._netResize&&window._netResize();}}}}
}}));

Chart.defaults.color='#617d9b';Chart.defaults.font.family="'JetBrains Mono',monospace";Chart.defaults.font.size=10;
const GRID={{color:'rgba(30,52,80,.5)'}};
const catNums=Object.keys(CATS).map(Number);

// Isolate-on-click legend plugin
function isolateLegend(chart){{
  const orig=Chart.defaults.plugins.legend.onClick;
  chart.options.plugins.legend.onClick=function(e,item,legend){{
    const idx=item.datasetIndex;
    const allMetas=chart.data.datasets.map((_,i)=>chart.getDatasetMeta(i));
    const isOnlyVisible=allMetas.every((m,i)=>i===idx?!m.hidden:m.hidden);
    if(isOnlyVisible){{allMetas.forEach(m=>m.hidden=false);}}
    else{{allMetas.forEach((m,i)=>m.hidden=(i!==idx));}}
    chart.update();
  }};
}}

// ── Overview ──────────────────────────────────────────────────────────────────
const ovChart=new Chart(document.getElementById('chtOv'),{{type:'bar',
  data:{{labels:YEARS,datasets:catNums.map(cn=>({{\
    label:`${{cn}}. ${{CATS[cn].short}}`,
    data:YEARS.map(y=>(CAT_YEAR[cn]||{{}})[y]||0),
    backgroundColor:h2r(CATS[cn].color,.7),borderColor:CATS[cn].color,borderWidth:1,borderRadius:2
  }}))
  }},options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{position:'bottom',labels:{{boxWidth:9,padding:9,color:'#617d9b',font:{{size:9}}}}}},
      tooltip:{{backgroundColor:'#0f1d2e',borderColor:'#1e3450',borderWidth:1}}}},
    scales:{{x:{{stacked:true,grid:GRID}},y:{{stacked:true,grid:GRID,min:0}}}}
  }}
}});
isolateLegend(ovChart);

const sts=['Oral Session','Symposium','ePoster Flash','Educational','Meet-the-Expert','Keynote','Late-Breaking','Pipeline','Case Session','IS / Integrated','Other'];
const stC=['#4fc3f7','#818cf8','#fbbf24','#34d399','#f472b6','#fb923c','#f87171','#a3e635','#6ee7b7','#e879f9','#94a3b8'];
const typeCounts={{}};
ROWS.forEach(r=>{{if(!typeCounts[r.type])typeCounts[r.type]={{}};typeCounts[r.type][r.year]=(typeCounts[r.type][r.year]||0)+1;}});
const tyChart=new Chart(document.getElementById('chtTy'),{{type:'bar',
  data:{{labels:YEARS,datasets:sts.map((s,i)=>({{\
    label:s,data:YEARS.map(y=>(typeCounts[s]||{{}})[y]||0),
    backgroundColor:stC[i]+'cc',borderColor:stC[i],borderWidth:1,borderRadius:2
  }}))
  }},options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{position:'bottom',labels:{{boxWidth:9,padding:8,color:'#617d9b',font:{{size:9}}}}}},
      tooltip:{{backgroundColor:'#0f1d2e',borderColor:'#1e3450',borderWidth:1}}}},
    scales:{{x:{{stacked:true,grid:GRID}},y:{{stacked:true,grid:GRID,min:0}}}}
  }}
}});

// ── Trends ────────────────────────────────────────────────────────────────────
const allChart=new Chart(document.getElementById('chtAll'),{{type:'line',
  data:{{labels:YEARS,datasets:catNums.map(cn=>({{\
    label:`${{cn}}. ${{CATS[cn].short}}`,
    data:YEARS.map(y=>(CAT_YEAR[cn]||{{}})[y]||0),
    borderColor:CATS[cn].color,backgroundColor:h2r(CATS[cn].color,.06),
    borderWidth:2,pointRadius:4,tension:.3,fill:true
  }}))
  }},options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{position:'bottom',labels:{{boxWidth:9,padding:8,color:'#617d9b',font:{{size:9}}}}}},
      tooltip:{{backgroundColor:'#0f1d2e',borderColor:'#1e3450',borderWidth:1}}}},
    scales:{{x:{{grid:GRID}},y:{{grid:GRID,min:0}}}}
  }}
}});
isolateLegend(allChart);

// Per-category subcategory trend charts
const trendGrid=document.getElementById('cat-trend-grid');
catNums.forEach(cn=>{{
  const cat=CATS[cn];
  const subcats=Object.entries(cat.subcats);
  // Build subcat×year counts from rows
  const scYear={{}};
  ROWS.filter(r=>r.cat===cn).forEach(r=>{{
    if(r.subcat){{if(!scYear[r.subcat])scYear[r.subcat]={{}};scYear[r.subcat][r.year]=(scYear[r.subcat][r.year]||0)+1;}}
  }});
  const hasSubs=subcats.some(([sc])=>YEARS.some(y=>(scYear[sc]||{{}})[y]));
  if(!hasSubs)return;
  
  const div=document.createElement('div');div.className='cat-trend-card';
  div.innerHTML=`<div class="cat-trend-title"><div class="cat-dot" style="background:${{cat.color}}"></div>${{cn}}. ${{cat.name}}</div><div style="position:relative;height:160px"><canvas id="ct_${{cn}}"></canvas></div>`;
  trendGrid.appendChild(div);
  
  // Pastel shades of category colour for subcats
  const subColors=['#fff','#ddd','#bbb','#999','#777','#555','#333','#111'].map((_,i)=>h2r(cat.color,0.4+i*0.08));
  new Chart(document.getElementById(`ct_${{cn}}`),{{type:'line',
    data:{{labels:YEARS,datasets:subcats.map(([sc,sn],i)=>({{\
      label:sn,data:YEARS.map(y=>(scYear[sc]||{{}})[y]||0),
      borderColor:subColors[i]||cat.color,backgroundColor:'transparent',
      borderWidth:1.5,pointRadius:3,tension:.3
    }}))
    }},options:{{responsive:true,maintainAspectRatio:false,
      plugins:{{legend:{{position:'bottom',labels:{{boxWidth:8,padding:6,color:'#617d9b',font:{{size:8}}}}}},
        tooltip:{{backgroundColor:'#0f1d2e',borderColor:'#1e3450',borderWidth:1}}}},
      scales:{{x:{{grid:GRID,ticks:{{font:{{size:8}}}}}},y:{{grid:GRID,min:0,ticks:{{font:{{size:8}}}}}}}}
    }}
  }});
}});

// ── Heatmaps ──────────────────────────────────────────────────────────────────
const tt=document.getElementById('tt');
function makeHeatmap(containerId,labels,colorFn,valueFn,tooltipFn){{
  const hmc=document.getElementById(containerId);
  const nYears=YEARS.length;
  hmc.style.cssText=`display:grid;grid-template-columns:180px repeat(${{nYears}},58px);gap:2px`;
  hmc.innerHTML='<div></div>'+YEARS.map(y=>`<div class="hm-yr">${{y}}</div>`).join('');
  const maxV=Math.max(...labels.map(l=>Math.max(...YEARS.map(y=>valueFn(l,y)))));
  labels.forEach(lbl=>{{
    hmc.innerHTML+=`<div class="hm-lbl" title="${{lbl}}">${{lbl.length>28?lbl.slice(0,27)+'…':lbl}}</div>`;
    YEARS.forEach(y=>{{
      const v=valueFn(lbl,y);
      const a=v===0?0:0.08+(v/maxV)*0.8;
      const col=colorFn(lbl);
      const c=document.createElement('div');c.className='hm-cell';c.textContent=v||'';
      c.style.background=v>0?h2r(col,a):'rgba(255,255,255,.03)';
      if(v>0)c.style.color=a>0.5?'#fff':'#c8d8ec';
      c.addEventListener('mousemove',e=>{{tt.className='tt on';tt.innerHTML=tooltipFn(lbl,y,v);tt.style.left=(e.clientX+12)+'px';tt.style.top=(e.clientY-36)+'px'}});
      c.addEventListener('mouseleave',()=>tt.className='tt');
      hmc.appendChild(c);
    }});
  }});
}}

// Left heatmap: subcategories
const subcatLabels=[];
catNums.forEach(cn=>Object.entries(CATS[cn].subcats).forEach(([sc,sn])=>subcatLabels.push(sc+'|'+cn+'|'+sn)));
const subcatCounts={{}};
ROWS.forEach(r=>{{if(r.subcat)subcatCounts[r.subcat+'|'+r.cat+'|'+(CATS[r.cat]||{{}}).subcats?.[r.subcat]]={{...subcatCounts[r.subcat+'|'+r.cat+'|'+(CATS[r.cat]||{{}}).subcats?.[r.subcat]]||{{}},
  [r.year]:((subcatCounts[r.subcat+'|'+r.cat+'|'+(CATS[r.cat]||{{}}).subcats?.[r.subcat]]||{{}})[r.year]||0)+1}};}});

const subYr={{}};
ROWS.forEach(r=>{{if(r.subcat){{const k=r.subcat;if(!subYr[k])subYr[k]={{}};subYr[k][r.year]=(subYr[k][r.year]||0)+1;}}  }});
// Sort subcats by total descending
const sortedSubcats=subcatLabels.filter(lbl=>{{const sc=lbl.split('|')[0];return YEARS.some(y=>(subYr[sc]||{{}})[y]>0);}})
  .sort((a,b)=>{{const sa=Object.values(subYr[a.split('|')[0]]||{{}}).reduce((x,y)=>x+y,0);const sb=Object.values(subYr[b.split('|')[0]]||{{}}).reduce((x,y)=>x+y,0);return sb-sa;}});

makeHeatmap('hmc-l',sortedSubcats,
  lbl=>catColor(parseInt(lbl.split('|')[1])),
  (lbl,y)=>(subYr[lbl.split('|')[0]]||{{}})[y]||0,
  (lbl,y,v)=>{{const[sc,cn,sn]=lbl.split('|');return `<strong style="color:${{catColor(parseInt(cn))}}">${{y}}</strong> · ${{sn||sc}}<br><span style="font-family:JetBrains Mono,monospace;font-size:13px;color:#e8f4ff">${{v}}</span>`}}
);

// Right heatmap: tags
const tagLabels=Object.values(TAG_TAX).flat();
const sortedTags=[...tagLabels].filter(t=>Object.keys(TAG_YEAR).includes(t))
  .sort((a,b)=>Object.values(TAG_YEAR[b]||{{}}).reduce((x,y)=>x+y,0)-Object.values(TAG_YEAR[a]||{{}}).reduce((x,y)=>x+y,0));
const untagged=tagLabels.filter(t=>!Object.keys(TAG_YEAR).includes(t));
const allTagsSorted=[...sortedTags,...untagged];

makeHeatmap('hmc-r',allTagsSorted,
  lbl=>tagColor(lbl),
  (lbl,y)=>(TAG_YEAR[lbl]||{{}})[y]||0,
  (lbl,y,v)=>`<strong style="color:${{tagColor(lbl)}}">${{y}}</strong> · ${{lbl}}<br><span style="font-family:JetBrains Mono,monospace;font-size:13px;color:#e8f4ff">${{v}}</span>`
);

// ── Network ────────────────────────────────────────────────────────────────────
function buildNetwork(){{
  const wrap=document.getElementById('net-svg-wrap');
  let W=wrap.getBoundingClientRect().width||900;
  const ctrlH=document.getElementById('net-ctrl').offsetHeight||44;
  const cardH=document.getElementById('net-card').offsetHeight||520;
  let H=cardH-ctrlH;if(H<380)H=460;

  const svg=d3.select('#net-svg').attr('width',W).attr('height',H);
  const mainG=svg.append('g');
  const zoom=d3.zoom().scaleExtent([.1,10]).on('zoom',e=>mainG.attr('transform',e.transform));
  svg.call(zoom);

  const nodes=D.net_rows.map(r=>Object.assign({{}},r,{{x:W/2+(Math.random()-.5)*260,y:H/2+(Math.random()-.5)*260}}));
  const nodeById=Object.fromEntries(nodes.map(n=>[n.net_id!==undefined?n.net_id:n.id,n]));
  const links=D.net_edges.map(e=>{{const s=nodeById[e.s],t=nodeById[e.t];return s&&t?{{source:s,target:t,w:e.w}}:null;}}).filter(Boolean);

  // Cluster centres — 12 categories in circle
  const cx=W/2,cy=H/2,cr=Math.min(W,H)*.31;
  const clusterC={{}};
  catNums.forEach((cn,i)=>{{const a=(i/12)*2*Math.PI-Math.PI/2;clusterC[cn]={{x:cx+cr*Math.cos(a),y:cy+cr*Math.sin(a)}}}});

  catNums.forEach(cn=>{{
    const c=clusterC[cn];
    mainG.append('circle').attr('cx',c.x).attr('cy',c.y).attr('r',Math.min(W,H)*.1)
      .attr('fill',h2r(catColor(cn),.03)).attr('stroke',h2r(catColor(cn),.14))
      .attr('stroke-width',1).attr('stroke-dasharray','4,3');
    mainG.append('text').attr('x',c.x).attr('y',c.y-Math.min(W,H)*.105)
      .attr('text-anchor','middle').attr('fill',h2r(catColor(cn),.22))
      .style('font-family','Sora,sans-serif').style('font-size','9px')
      .style('font-weight','700').style('letter-spacing','.07em').style('text-transform','uppercase')
      .text(`${{cn}}. ${{(CATS[cn]||{{}}).short||''}}`);
  }});

  const linkSel=mainG.append('g').selectAll('line').data(links).enter().append('line')
    .attr('class','lnk').attr('stroke',d=>catColor(d.source.cat))
    .attr('stroke-width',d=>Math.max(.5,d.w*.8));

  const nodeSel=mainG.append('g').selectAll('g.nd').data(nodes).enter().append('g').attr('class','nd')
    .call(d3.drag().on('start',ds).on('drag',dd).on('end',de));

  nodeSel.append('circle').attr('r',8.5).attr('fill','none')
    .attr('stroke',d=>YCOLORS[d.year]||'#617d9b').attr('stroke-width',1.2).attr('opacity',.4);
  nodeSel.append('circle').attr('class','nc').attr('r',5.5)
    .attr('fill',d=>h2r(catColor(d.cat),.9)).attr('stroke',d=>catColor(d.cat))
    .attr('stroke-width',1.5).style('cursor','pointer');

  function ds(e,d){{if(!e.active)sim.alphaTarget(.3).restart();d.fx=d.x;d.fy=d.y}}
  function dd(e,d){{d.fx=e.x;d.fy=e.y}}
  function de(e,d){{if(!e.active)sim.alphaTarget(0);d.fx=null;d.fy=null}}

  nodeSel.on('mousemove',function(e,d){{
    const tstr=(d.tags||[]).map(t=>t.replace(/^[a-z_]+_/,'')).join(' · ')||'(no tags yet)';
    tt.className='tt on';
    tt.innerHTML=`<strong style="color:${{catColor(d.cat)}}">${{d.year}} · ${{d.cat_short||''}}</strong><br>${{d.title.substring(0,75)}}…<br><small style="color:var(--dim)">${{tstr}}</small>`;
    tt.style.left=(e.clientX+14)+'px';tt.style.top=(e.clientY-44)+'px';
  }}).on('mouseleave',()=>tt.className='tt');

  let selId=null;
  nodeSel.on('click',function(e,d){{e.stopPropagation();selectNode(d)}});
  svg.on('click',()=>selectNode(null));

  function selectNode(d){{
    selId=d?d.id:null;
    const panel=document.getElementById('dp');
    if(!d){{
      nodeSel.attr('opacity',1);linkSel.attr('stroke-opacity',.1);
      panel.innerHTML='<div class="de">Click any node<br>to see details</div>';return;
    }}
    const nbIds=new Set();
    links.forEach(l=>{{if(l.source.id===d.id)nbIds.add(l.target.id);if(l.target.id===d.id)nbIds.add(l.source.id)}});
    const connIdx=new Set();links.forEach((l,i)=>{{if(l.source.id===d.id||l.target.id===d.id)connIdx.add(i)}});
    nodeSel.attr('opacity',n=>n.id===d.id||nbIds.has(n.id)?1:.06);
    linkSel.attr('stroke-opacity',(_,i)=>connIdx.has(i)?.6:.01);
    const tagpills=(d.tags||[]).map(t=>`<span class="dtag" style="background:${{h2r(tagColor(t),.15)}};color:${{tagColor(t)}};border:1px solid ${{h2r(tagColor(t),.3)}}">${{t.replace(/^[a-z_]+_/,'')}}</span>`).join('');
    const catpill=`<span class="dtag" style="background:${{h2r(catColor(d.cat),.15)}};color:${{catColor(d.cat)}};border:1px solid ${{h2r(catColor(d.cat),.3)}}">${{d.cat}}. ${{d.cat_short||''}}</span>`;
    const nbChips=[...nbIds].slice(0,5).map(nid=>{{const nb=nodeById[nid];return nb?`<button class="nbc" onclick="window._sel(${{nid}})">${{nb.year}}: ${{nb.title.substring(0,50)}}…</button>`:''}}).join('');
    panel.innerHTML=`<div class="dy" style="color:${{catColor(d.cat)}}">${{d.year}} · ${{d.type}}</div><div class="dt">${{d.title}}</div><div class="dtags">${{catpill}}${{tagpills||'<span style="color:var(--dim);font-size:10px">No tags yet</span>'}}</div>${{nbIds.size?`<div class="dnbl">Connected (${{nbIds.size}})</div>${{nbChips}}${{nbIds.size>5?`<div style="font-size:10px;color:var(--dim)">+${{nbIds.size-5}} more…</div>`:''}}`:''}}`;
  }}
  window._sel=id=>selectNode(nodeById[id]);

  let clusterOn=true,edgesOn=true;
  function clusterForce(alpha){{if(!clusterOn)return;nodes.forEach(n=>{{const c=clusterC[n.cat];if(!c)return;n.vx-=(n.x-c.x)*.04*alpha;n.vy-=(n.y-c.y)*.04*alpha}})}}
  const sim=d3.forceSimulation(nodes)
    .force('link',d3.forceLink(links).id(d=>d.id).distance(35).strength(d=>Math.min(d.w*.2,.6)))
    .force('charge',d3.forceManyBody().strength(-70))
    .force('collide',d3.forceCollide(10).strength(.8))
    .force('cluster',clusterForce)
    .force('center',d3.forceCenter(W/2,H/2).strength(.03))
    .on('tick',()=>{{
      linkSel.attr('x1',d=>d.source.x).attr('y1',d=>d.source.y).attr('x2',d=>d.target.x).attr('y2',d=>d.target.y);
      nodeSel.attr('transform',d=>`translate(${{d.x}},${{d.y}})`);
    }});

  const yrEl=document.getElementById('net-yr');
  let activeYrs=new Set(YEARS);
  YEARS.forEach(y=>{{
    const b=document.createElement('button');b.className='cb on';
    b.style.cssText=`background:${{h2r(YCOLORS[y],.22)}};border-color:${{YCOLORS[y]}}66;color:${{YCOLORS[y]}}`;
    b.textContent=y;
    b.onclick=()=>{{activeYrs.has(y)?(activeYrs.size>1&&(activeYrs.delete(y),b.classList.remove('on'),b.style.background='transparent')):(activeYrs.add(y),b.classList.add('on'),b.style.background=h2r(YCOLORS[y],.22));applyVis()}};
    yrEl.appendChild(b);
  }});

  const catEl=document.getElementById('net-cat');
  let activeCats=new Set(catNums);
  catNums.forEach(cn=>{{
    const b=document.createElement('button');b.className='cb on';
    b.style.cssText=`background:${{h2r(catColor(cn),.22)}};border-color:${{catColor(cn)}}66;color:${{catColor(cn)}}`;
    b.textContent=`${{cn}}`;b.title=(CATS[cn]||{{}}).name;
    b.onclick=()=>{{activeCats.has(cn)?(activeCats.size>1&&(activeCats.delete(cn),b.classList.remove('on'),b.style.background='transparent')):(activeCats.add(cn),b.classList.add('on'),b.style.background=h2r(catColor(cn),.22));applyVis()}};
    catEl.appendChild(b);
  }});

  function applyVis(){{
    nodeSel.attr('opacity',d=>activeYrs.has(d.year)&&activeCats.has(d.cat)?1:.03);
    linkSel.attr('stroke-opacity',d=>{{const vs=activeYrs.has(d.source.year)&&activeCats.has(d.source.cat);const vt=activeYrs.has(d.target.year)&&activeCats.has(d.target.cat);return vs&&vt&&edgesOn?.1:0;}});
  }}
  document.getElementById('btn-edges').onclick=function(){{edgesOn=!edgesOn;this.textContent=edgesOn?'Edges on':'Edges off';this.classList.toggle('on',edgesOn);linkSel.attr('stroke-opacity',edgesOn?.1:0)}};
  document.getElementById('btn-cluster').onclick=function(){{clusterOn=!clusterOn;this.textContent=clusterOn?'Cluster on':'Cluster off';this.classList.toggle('on',clusterOn);if(clusterOn)sim.alpha(.4).restart()}};
  document.getElementById('btn-reset').onclick=()=>{{svg.transition().duration(500).call(zoom.transform,d3.zoomIdentity);selectNode(null)}};
  window._netResize=()=>{{const nW=wrap.getBoundingClientRect().width;if(nW>50){{W=nW;svg.attr('width',W);sim.force('center',d3.forceCenter(W/2,H/2).strength(.03)).alpha(.2).restart()}}}};
  new ResizeObserver(()=>{{if(netInit)window._netResize&&window._netResize()}}).observe(wrap);
}}

// ── Explore ───────────────────────────────────────────────────────────────────
const HIDE_TYPES=new Set(["ePoster Flash","IS / Integrated","Other"]);
const sorted=[...ROWS].sort((a,b)=>b.year.localeCompare(a.year)||a.title.localeCompare(b.title));
let eYear='',eCat='',eSub='',eTag='',eSearch='',eShowAll=false;

// ePoster toggle
document.getElementById('btn-show-all').onclick=function(){{
  eShowAll=!eShowAll;
  this.textContent=eShowAll?'All types shown':'ePosters hidden';
  this.classList.toggle('on',eShowAll);
  render();
}};

// Populate year dropdown
const yrSel=document.getElementById('ef-yr-sel');
YEARS.slice().reverse().forEach(y=>{{yrSel.innerHTML+=`<option value="${{y}}">${{y}}</option>`;}});
yrSel.onchange=()=>{{eYear=yrSel.value;refreshSubcats();refreshTags();render()}};

// Populate category dropdown
const catSel=document.getElementById('ef-cat-sel');
catNums.forEach(cn=>{{catSel.innerHTML+=`<option value="${{cn}}">${{cn}}. ${{(CATS[cn]||{{}}).name||''}}</option>`;}});
catSel.onchange=()=>{{
  eCat=catSel.value; eSub=''; eTag='';
  subSel.value=''; tagSel.value='';
  refreshSubcats(); refreshTags(); render();
}};

// Subcategory dropdown — all subcats always available; selecting one auto-sets category
const subSel=document.getElementById('ef-sub-sel');
function refreshSubcats(){{
  const cn=eCat?parseInt(eCat):null;
  // Build counts across current year filter
  const scCounts={{}};
  sorted.filter(r=>(!eYear||r.year===eYear)&&(!cn||r.cat===cn))
    .forEach(r=>{{if(r.subcat)scCounts[r.subcat]=(scCounts[r.subcat]||0)+1;}});
  const prev=subSel.value;
  subSel.innerHTML='<option value="">All subcategories</option>';
  // Group by category
  catNums.forEach(cn2=>{{
    const cat=CATS[cn2]||{{}};
    if(cn&&cn2!==cn)return; // filtered to one cat
    Object.entries(cat.subcats||{{}}).forEach(([sc,sn])=>{{
      const n=scCounts[sc]||0;
      if(n>0)subSel.innerHTML+=`<option value="${{sc}}">${{sc}}: ${{sn}} (${{n}})</option>`;
    }});
  }});
  subSel.value=prev&&subSel.querySelector(`option[value="${{prev}}"]`)?prev:'';
}}
subSel.onchange=()=>{{
  eSub=subSel.value;
  // Auto-set category to match selected subcategory
  if(eSub){{
    for(const cn2 of catNums){{
      if(Object.keys((CATS[cn2]||{{}}).subcats||{{}}).includes(eSub)){{
        eCat=String(cn2); catSel.value=eCat; break;
      }}
    }}
  }}
  refreshTags(); render();
}};

// Tag dropdown — all tags always available; sorted by frequency
const tagSel=document.getElementById('ef-tag-sel');
function refreshTags(){{
  const cn=eCat?parseInt(eCat):null;
  const tagCounts={{}};
  sorted.filter(r=>(!eYear||r.year===eYear)&&(!cn||r.cat===cn)&&(!eSub||r.subcat===eSub))
    .forEach(r=>(r.tags||[]).forEach(t=>tagCounts[t]=(tagCounts[t]||0)+1));
  const prev=tagSel.value;
  tagSel.innerHTML='<option value="">All tags</option>';
  Object.entries(tagCounts).sort((a,b)=>b[1]-a[1])
    .forEach(([t,n])=>{{tagSel.innerHTML+=`<option value="${{t}}">${{t.replace(/^[a-z_]+_/,'')}} (${{n}})</option>`;}});
  tagSel.value=prev&&tagSel.querySelector(`option[value="${{prev}}"]`)?prev:'';
}}
tagSel.onchange=()=>{{eTag=tagSel.value;render()}};

refreshSubcats();
refreshTags();

function render(){{
  const q=eSearch.toLowerCase();
  const cn=eCat?parseInt(eCat):null;
  const f=sorted.filter(r=>{{
    if(!eShowAll&&HIDE_TYPES.has(r.type))return false;
    if(eYear&&r.year!==eYear)return false;
    if(cn&&r.cat!==cn)return false;
    if(eSub&&r.subcat!==eSub)return false;
    if(eTag&&!(r.tags||[]).includes(eTag))return false;
    if(q&&!r.title.toLowerCase().includes(q))return false;
    return true;
  }});
  document.getElementById('etc').textContent=f.length.toLocaleString()+' sessions';
  document.getElementById('tlist').innerHTML=f.map(r=>{{
    const cat=CATS[r.cat]||{{}};
    const catpill=`<span class="tp" style="background:${{h2r(cat.color||'#617d9b',.15)}};color:${{cat.color||'#617d9b'}};border:1px solid ${{h2r(cat.color||'#617d9b',.3)}}">${{r.cat}}.${{cat.short||''}}</span>`;
    const tagpills=(r.tags||[]).map(t=>`<span class="tp" style="background:${{h2r(tagColor(t),.15)}};color:${{tagColor(t)}};border:1px solid ${{h2r(tagColor(t),.3)}}">${{t.replace(/^[a-z_]+_/,'')}}</span>`).join('');
    return `<div class="ti" style="cursor:pointer" onclick="openModal(${{r.id}})"><div class="ty">${{r.year}}</div><div class="tt2">${{r.title}}<div class="ts">${{r.type}}</div></div><div class="tgs">${{catpill}}${{tagpills}}</div></div>`;
  }}).join('');
}}
render();
document.getElementById('search').addEventListener('input',e=>{{eSearch=e.target.value;render()}});

// ── Session detail modal ──────────────────────────────────────────────────────
const rowById=Object.fromEntries(ROWS.map(r=>[r.id,r]));

function openModal(id){{
  const r=rowById[id];
  if(!r)return;
  const cat=CATS[r.cat]||{{}};
  const cat2=r.cat2?CATS[r.cat2]:null;
  const tagpills=(r.tags||[]).map(t=>`<span style="display:inline-block;margin:2px 3px 2px 0;padding:2px 9px;border-radius:12px;font-size:10px;font-family:'JetBrains Mono',monospace;background:${{h2r(tagColor(t),.18)}};color:${{tagColor(t)}};border:1px solid ${{h2r(tagColor(t),.3)}}">${{t.replace(/^[a-z_]+_/,'')}}</span>`).join('');
  const chairHtml=r.chairs?`<div style="margin-top:14px"><div style="font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:.1em;color:var(--dim);margin-bottom:4px">Chairs</div><div style="font-size:12px;color:var(--text)">${{r.chairs}}</div></div>`:'';
  const timeHtml=r.time_hall?`<div style="font-size:11px;color:var(--dim);margin-top:4px;font-family:'JetBrains Mono',monospace">${{r.time_hall}}</div>`:'';
  const noTagNote=!(r.tags&&r.tags.length)?`<div style="font-size:11px;color:var(--dim);margin-top:10px;font-style:italic">Tags not yet assigned — run <code style="font-size:10px;background:var(--surf2);padding:1px 5px;border-radius:3px">python escmid_analyser.py --add-tags</code> to populate</div>`:'';
  // Talks section
  const talkRows=(r.talks||[]).map(t=>{{
    const spk=t.speakers?`<div style="font-size:10px;color:var(--dim);margin-top:2px">↳ ${{t.speakers.substring(0,120)}}</div>`:'';
    return `<div style="padding:7px 0;border-bottom:1px solid var(--bdr)"><div style="display:flex;gap:8px;align-items:baseline"><span style="font-family:'JetBrains Mono',monospace;font-size:10px;color:var(--acc);flex-shrink:0">${{t.code}}${{t.time?' · '+t.time:''}}</span><span style="font-size:12px;color:var(--text)">${{t.title}}</span></div>${{spk}}</div>`;
  }}).join('');
  const talksHtml=talkRows?`<div style="margin-top:16px"><div style="font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:.1em;color:var(--dim);margin-bottom:8px">Individual talks (${{(r.talks||[]).length}})</div>${{talkRows}}</div>`
    :`<div style="margin-top:14px;padding:10px 14px;background:var(--surf2);border:1px solid var(--bdr);border-radius:6px;font-size:11px;color:var(--dim)">💡 Re-run with <code style="font-size:10px;background:var(--surf);padding:1px 5px;border-radius:3px">--extract-talks</code> to surface individual abstract titles and speaker names.</div>`;
  const absNote='';
  document.getElementById('modal-body').innerHTML=`
    <div style="font-size:11px;color:var(--dim);margin-bottom:8px;font-family:'JetBrains Mono',monospace">${{r.year}} · ${{r.code}} · ${{r.type}}</div>
    <div style="font-size:16px;font-weight:700;color:#e8f4ff;line-height:1.4;margin-bottom:12px">${{r.title}}</div>
    ${{timeHtml}}
    <div style="margin-top:12px;display:flex;flex-wrap:wrap;gap:8px">
      <span style="padding:4px 12px;border-radius:14px;font-size:11px;font-weight:600;background:${{h2r(cat.color||'#617d9b',.2)}};color:${{cat.color||'#617d9b'}};border:1px solid ${{h2r(cat.color||'#617d9b',.4)}}">${{r.cat}}. ${{cat.name||''}}</span>
      ${{r.subcat?`<span style="padding:4px 12px;border-radius:14px;font-size:11px;background:rgba(255,255,255,.06);color:var(--dim)">${{r.subcat}}: ${{r.subcat_name||''}}</span>`:''}}</div>
      ${{cat2?`<div style="font-size:11px;color:var(--dim);margin-top:6px">Also: ${{r.cat2}}. ${{cat2.name}}</div>`:''}}
    ${{chairHtml}}
    ${{r.tags&&r.tags.length?`<div style="margin-top:14px"><div style="font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:.1em;color:var(--dim);margin-bottom:6px">Tags</div>${{tagpills}}</div>`:noTagNote}}
    ${{talksHtml}}
  `;
  const ov=document.getElementById('modal-overlay');
  ov.style.display='flex';
  setTimeout(()=>ov.style.opacity=1,10);
}}
function closeModal(){{
  document.getElementById('modal-overlay').style.display='none';
}}
document.addEventListener('keydown',e=>{{if(e.key==='Escape')closeModal();}});

// ── People ────────────────────────────────────────────────────────────────────
const PEOPLE = D.people || [];
let pSearch='', pYear='', pRole='', pCat='';

// Populate People filters
const pyEl=document.getElementById('p-yr');
YEARS.forEach(y=>{{pyEl.innerHTML+=`<option value="${{y}}">${{y}}</option>`;}});
pyEl.onchange=()=>{{pYear=pyEl.value;renderPeople()}};

const prEl=document.getElementById('p-role');
prEl.onchange=()=>{{pRole=prEl.value;renderPeople()}};

const pcEl=document.getElementById('p-cat');
catNums.forEach(cn=>{{pcEl.innerHTML+=`<option value="${{cn}}">${{cn}}. ${{(CATS[cn]||{{}}).short||''}}</option>`;}});
pcEl.onchange=()=>{{pCat=pcEl.value;renderPeople()}};

document.getElementById('p-search').addEventListener('input',e=>{{pSearch=e.target.value.toLowerCase();renderPeople()}});

function renderPeople(){{
  const cn=pCat?parseInt(pCat):null;
  const f=PEOPLE.filter(p=>{{
    if(pSearch&&!p.name.toLowerCase().includes(pSearch))return false;
    if(pYear&&!p.years.includes(pYear))return false;
    if(pRole==='Chair'&&!p.chair_count)return false;
    if(pRole==='Speaker'&&!p.speaker_count)return false;
    if(cn&&!p.apps.some(a=>a.cat===cn))return false;
    return true;
  }});
  document.getElementById('p-count').textContent=f.length.toLocaleString()+' people';
  document.getElementById('p-list').innerHTML=f.map((p,i)=>{{
    const yearPills=p.years.map(y=>`<span style="font-family:'JetBrains Mono',monospace;font-size:9px;padding:2px 7px;border-radius:10px;background:${{h2r(YCOLORS[y]||'#617d9b',.25)}};color:${{YCOLORS[y]||'#617d9b'}}">${{y}}</span>`).join('');
    const roleStr=(p.chair_count&&p.speaker_count)?`Chair (${{p.chair_count}}) + Speaker (${{p.speaker_count}})`:p.chair_count?`Chair (${{p.chair_count}})`:`Speaker (${{p.speaker_count}})`;
    return `<div class="ti" style="cursor:pointer;grid-template-columns:1fr auto auto" onclick="openPersonModal(${{i}})">
      <div class="tt2" style="font-weight:600">${{p.name}}</div>
      <div style="display:flex;gap:4px;flex-wrap:wrap;justify-content:flex-end;max-width:300px">${{yearPills}}</div>
      <div style="font-size:10px;color:var(--dim);text-align:right;min-width:160px;padding-left:8px">${{roleStr}}</div>
    </div>`;
  }}).join('');
}}
renderPeople();

function openPersonModal(idx){{
  const f=PEOPLE.filter(p=>{{
    const cn=pCat?parseInt(pCat):null;
    if(pSearch&&!p.name.toLowerCase().includes(pSearch))return false;
    if(pYear&&!p.years.includes(pYear))return false;
    if(pRole==='Chair'&&!p.chair_count)return false;
    if(pRole==='Speaker'&&!p.speaker_count)return false;
    if(cn&&!p.apps.some(a=>a.cat===cn))return false;
    return true;
  }});
  const p=f[idx]; if(!p)return;
  const yearPills=p.years.map(y=>`<span style="font-family:'JetBrains Mono',monospace;font-size:10px;padding:3px 9px;border-radius:12px;background:${{h2r(YCOLORS[y]||'#617d9b',.25)}};color:${{YCOLORS[y]||'#617d9b'}}">${{y}}</span>`).join(' ');
  const appRows=p.apps.map(a=>{{
    const cat=CATS[a.cat]||{{}};
    const catpill=`<span style="font-size:9px;padding:1px 7px;border-radius:10px;background:${{h2r(a.color||'#617d9b',.15)}};color:${{a.color||'#617d9b'}};border:1px solid ${{h2r(a.color||'#617d9b',.3)}}">${{a.cat}}. ${{cat.short||''}}</span>`;
    const roleColor=a.role==='Chair'?'#fbbf24':'#4fc3f7';
    const talkLine=a.talk_code?`<div style="font-size:10px;color:var(--dim);margin-top:2px">${{a.talk_code}}: ${{a.talk_title}}</div>`:'';
    return `<div style="padding:8px 0;border-bottom:1px solid var(--bdr);display:flex;gap:10px;align-items:baseline">
      <span style="font-family:'JetBrains Mono',monospace;font-size:10px;color:${{YCOLORS[a.year]||'#617d9b'}};flex-shrink:0;min-width:38px">${{a.year}}</span>
      <span style="font-family:'JetBrains Mono',monospace;font-size:10px;color:var(--acc);flex-shrink:0;min-width:55px">${{a.code}}</span>
      <span style="font-size:9px;padding:1px 7px;border-radius:10px;background:${{h2r(roleColor,.15)}};color:${{roleColor}};flex-shrink:0">${{a.role}}</span>
      <div><div style="font-size:12px;color:var(--text)">${{a.title}}</div>${{talkLine}}</div>
      <div style="margin-left:auto;flex-shrink:0">${{catpill}}</div>
    </div>`;
  }}).join('');
  document.getElementById('person-body').innerHTML=`
    <div style="font-size:18px;font-weight:700;color:#e8f4ff;margin-bottom:10px">${{p.name}}</div>
    <div style="display:flex;gap:6px;flex-wrap:wrap;margin-bottom:14px">${{yearPills}}</div>
    <div style="font-size:12px;color:var(--dim);margin-bottom:16px">${{p.chair_count?p.chair_count+' session(s) as Chair':''}}&nbsp;${{p.speaker_count?p.speaker_count+' talk(s) as Speaker':''}}</div>
    ${{appRows}}
  `;
  document.getElementById('person-modal').style.display='flex';
}}
function closePersonModal(){{document.getElementById('person-modal').style.display='none';}}
document.addEventListener('keydown',e=>{{if(e.key==='Escape'){{closeModal();closePersonModal();}}}});

// ── Gaps ──────────────────────────────────────────────────────────────────────
// Subcategory gaps
const subTot={{}};
ROWS.forEach(r=>{{if(r.subcat)subTot[r.subcat]=(subTot[r.subcat]||0)+1;}});
const ggs=document.getElementById('gg-sub');
const allSubcatsFlat=[];
catNums.forEach(cn=>Object.entries((CATS[cn]||{{}}).subcats||{{}}).forEach(([sc,sn])=>allSubcatsFlat.push([sc,sn,cn,subTot[sc]||0])));
allSubcatsFlat.sort((a,b)=>a[3]-b[3]).slice(0,16).forEach(([sc,sn,cn,count])=>{{
  const cls=count<=5?'low':'med';
  ggs.innerHTML+=`<div class="gc ${{cls}}"><div class="gct">${{sc}}: ${{sn}}</div><div class="gcn ${{cls}}">${{count}}</div><div class="gcx">in ${{(CATS[cn]||{{}}).name||''}}</div></div>`;
}});

// Tag gaps
const allTagsList=Object.values(TAG_TAX).flat();
const tagTot={{}};allTagsList.forEach(t=>tagTot[t]=Object.values(TAG_YEAR[t]||{{}}).reduce((a,b)=>a+b,0));
const ggt=document.getElementById('gg-tag');
allTagsList.map(t=>[t,tagTot[t]||0]).sort((a,b)=>a[1]-b[1]).slice(0,16).forEach(([tag,count])=>{{
  const cls=count<=2?'low':'med';
  ggt.innerHTML+=`<div class="gc ${{cls}}"><div class="gct">${{tag}}</div><div class="gcn ${{cls}}">${{count}}</div><div class="gcx">sessions tagged</div></div>`;
}});
</script>
</body>
</html>"""

def generate_dashboard(sessions: list, path: Path):
    data = build_dashboard_data(sessions)
    html = generate_dashboard_html(json.dumps(data))
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(html, encoding="utf-8")
    print(f"  HTML   → {path}")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    ap = argparse.ArgumentParser(description="ESCMID EPC Copilot")
    ap.add_argument("--years",        help="Comma-separated years, e.g. 2024,2025,2026")
    ap.add_argument("--skip-tagging", action="store_true", help="Use keyword fallback only (free, less accurate)")
    ap.add_argument("--extract-talks", action="store_true",
                    help="Also extract individual talk titles and speakers within sessions")
    ap.add_argument("--rebuild",      action="store_true",
                    help="Re-extract sessions with pdfplumber but keep categories/tags "
                         "from existing sessions_raw.json — no API calls needed")
    ap.add_argument("--talks-only",    action="store_true",
                    help="Load existing sessions_raw.json, add talks, regenerate outputs — no API calls")
    ap.add_argument("--add-tags",     action="store_true",
                    help="After categorisation, also assign up to 4 tags from the 94-tag clinical taxonomy")
    ap.add_argument("--api-extract",  action="store_true",
                    help="Use API for BOTH extraction clean-up AND classification (recommended)")
    ap.add_argument("--skip-excel",   action="store_true", help="Skip Excel output")
    ap.add_argument("--skip-dash",    action="store_true", help="Skip HTML dashboard")
    ap.add_argument("--programmes-dir", default=str(PROGRAMMES_DIR),
                    help=f"Directory containing PDFs (default: {PROGRAMMES_DIR})")
    ap.add_argument("--output-dir",   default=str(OUTPUT_DIR),
                    help=f"Output directory (default: {OUTPUT_DIR})")
    args = ap.parse_args()

    prog_dir = Path(args.programmes_dir)
    out_dir  = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    # Select years
    if args.years:
        selected = set(y.strip() for y in args.years.split(","))
    else:
        selected = set(PROGRAMME_FILES.keys())

    pdf_map = {
        yr: prog_dir / fname
        for yr, fname in PROGRAMME_FILES.items()
        if yr in selected
    }

    print("=" * 58)
    print(" ESCMID EPC Copilot")
    print("=" * 58)

    # ── Shortcut: --rebuild — fresh extraction, keep existing categories/tags ──
    if args.rebuild:
        raw_path = out_dir / "sessions_raw.json"
        if not raw_path.exists():
            print(f"Error: {raw_path} not found. Run a full pipeline first.")
            sys.exit(1)
        print("\nLoading categories and tags from tagging_cache.json …")
        cache_path = CACHE_FILE
        if not cache_path.exists():
            print(f"Error: {cache_path} not found.")
            sys.exit(1)
        cache = json.loads(cache_path.read_text())

        # Build category lookup: {(year, code): category_dict}
        # Cache keys: "extract|year|code|title[:80]" or fallback by "year|code" prefix
        def _cat_from_cache(year, code):
            prefix = f"extract|{year}|{code}|"
            for k, v in cache.items():
                if k.startswith(prefix):
                    return v
            return {}

        # Build tag lookup: {"tags94|year|code|...": [tags]}
        def _tags_from_cache(year, code):
            prefix = f"tags94|{year}|{code}|"
            for k, v in cache.items():
                if k.startswith(prefix):
                    return v if isinstance(v, list) else []
            return []

        ext_count = sum(1 for k in cache if k.startswith("extract|"))
        tag_count = sum(1 for k in cache if k.startswith("tags94|"))
        print(f"  {ext_count} category entries, {tag_count} tag entries in cache")

        print("\nRe-extracting with pdfplumber for fresh titles/talks/chairs …")
        new_sessions = extract_all(pdf_map, include_talks=True)
        print(f"  {len(new_sessions)} sessions extracted")

        # Merge: fresh titles/talks/chairs + cached categories/tags
        merged = []
        for s in new_sessions:
            cat_data = _cat_from_cache(s["year"], s["code"])
            if cat_data:
                cat_num = int(cat_data.get("primary_cat") or 4)
                cat     = CATEGORIES.get(cat_num, CATEGORIES[4])
                s["escmid_cat"]         = cat_num
                s["escmid_cat_name"]    = cat["name"]
                s["escmid_cat_short"]   = cat["short"]
                s["escmid_color"]       = cat["color"]
                s["escmid_subcat"]      = str(cat_data.get("primary_subcat") or "")
                s["escmid_subcat_name"] = cat["subcats"].get(
                    s["escmid_subcat"], "")
                s["escmid_cat2"]        = cat_data.get("secondary_cat")
                s["escmid_subcat2"]     = str(cat_data.get("secondary_subcat") or "")
                s["confidence"]         = str(cat_data.get("confidence") or "")
            else:
                s = enrich_session(s, keyword_tag(s))
            s["tags"] = _tags_from_cache(s["year"], s["code"])
            merged.append(s)

        from collections import Counter as _C
        cat_dist = _C(s["escmid_cat"] for s in merged)
        print("  Category distribution:")
        for cn, n in sorted(cat_dist.items()):
            print(f"    {cn:2d}. {CATEGORIES.get(cn,{}).get('short','?'):<22} {n:>4}")

        raw_path.write_text(json.dumps(merged, indent=2))
        print("\nGenerating outputs …")
        if not args.skip_excel:
            generate_excel(merged, out_dir / "ESCMID_Programmes.xlsx")
        if not args.skip_dash:
            generate_dashboard(merged, out_dir / "ESCMID_Dashboard.html")
        print(f"\nDone — outputs in {out_dir}")
        return

    # ── Shortcut: --talks-only skips extraction and tagging ───
    if args.talks_only:
        raw_path = out_dir / "sessions_raw.json"
        if not raw_path.exists():
            print(f"Error: {raw_path} not found. Run without --talks-only first.")
            sys.exit(1)
        print(f"\nLoading existing sessions from {raw_path} …")
        sessions = json.loads(raw_path.read_text())
        print(f"  {len(sessions)} sessions loaded")
        print("\nAdding talks from PDFs …")
        for year, path in sorted(pdf_map.items()):
            if not path.exists():
                continue
            print(f"  {year} …", end=" ", flush=True)
            try:
                text = pdf_to_text(path)
                talks_map = extract_talks_from_text(text)
                added = 0
                for s in sessions:
                    if s["year"] == year and s["code"] in talks_map:
                        s["talks"] = talks_map[s["code"]]
                        added += len(s["talks"])
                print(f"{added} talks")
            except Exception as e:
                print(f"ERROR — {e}")
        raw_path.write_text(json.dumps(sessions, indent=2))
        print("\nGenerating outputs …")
        if not args.skip_excel:
            generate_excel(sessions, out_dir / "ESCMID_Programmes.xlsx")
        if not args.skip_dash:
            generate_dashboard(sessions, out_dir / "ESCMID_Dashboard.html")
        print(f"\nDone — outputs in {out_dir}")
        return

    # ── 1. Extract ────────────────────────────────────────────
    print("\n[1/3] Extracting sessions from PDFs …")
    sessions = extract_all(pdf_map, include_talks=args.extract_talks)
    print(f"  Total: {len(sessions)} sessions across {len(pdf_map)} years")

    if not sessions:
        print("  No sessions found. Check PDF paths and try again.")
        sys.exit(1)

    # ── 2. Tag ────────────────────────────────────────────────
    print("\n[2/3] Tagging sessions …")
    use_api = not args.skip_tagging and not args.api_extract
    sessions = tag_all(sessions, use_api=use_api, api_extract=args.api_extract)

    # Optional: add 94-tag clinical taxonomy
    if args.add_tags:
        print("\n[2b/3] Adding 94-tag clinical taxonomy ...")
        sessions = run_tag_pass(sessions)

    # Save raw JSON (useful for debugging / re-running without re-tagging)
    raw_path = out_dir / "sessions_raw.json"
    raw_path.write_text(json.dumps(sessions, indent=2))
    print(f"  Raw data → {raw_path}")

    # Category counts summary
    counts = Counter(s["escmid_cat"] for s in sessions)
    print("  Category breakdown:")
    for cn in sorted(counts):
        print(f"    {cn:2d}. {CATEGORIES[cn]['short']:<22} {counts[cn]:>4} sessions")

    # ── 3. Output ─────────────────────────────────────────────
    print("\n[3/3] Generating outputs …")
    if not args.skip_excel:
        generate_excel(sessions, out_dir / "ESCMID_Programmes.xlsx")
    if not args.skip_dash:
        generate_dashboard(sessions, out_dir / "ESCMID_Dashboard.html")

    print(f"\nDone — outputs in {out_dir}")


if __name__ == "__main__":
    main()
